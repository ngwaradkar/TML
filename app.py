"""
TCF PPC Dashboard — Production Planning & Shortage Report
═════════════════════════════════════════════════════════════
Streamlit application for Tata Motors TCF line tracking.
Reads raw data uploads, computes shortage reports against
paint float pipeline, and presents 4 interactive summary tabs.
"""

import streamlit as st
import pandas as pd
import numpy as np

pd.set_option('future.no_silent_downcasting', True)
import os
import json
import re
import shutil
from pathlib import Path
from io import BytesIO
from datetime import datetime, time as dtime, timezone, timedelta

def get_ist_now():
    return datetime.now(timezone(timedelta(hours=5, minutes=30)))


# ═══════════════════════════════════════════════════════════════
# 1. CONSTANTS & CONFIGURATION
# ═══════════════════════════════════════════════════════════════

DATA_DIR = Path("data")
ENGINE_FILE = DATA_DIR / "engine_stock.json"
BOM_FILE = DATA_DIR / "bom_master.json"
RESET_STATE_FILE = DATA_DIR / "last_reset.json"
BOM_SOURCE_FILE = Path("d:/TML PPC Dashboard/Bom details.xlsx")

# Daily reset happens at 06:30 local time
RESET_HOUR = 6
RESET_MINUTE = 30

RAW_FILE_TYPES = [
    "Paint Float",
    "TCF1 Wiring File",
    "TCF2 Wiring File",
    "TCF1 Cockpit",
    "TCF2 Cockpit",
    "Nova Cockpit",
    "TCF1 DPT Plan",
    "TCF2 DPT Plan",
    "Engine Manual Data",
]

MODEL_MAP = {
    "5497": ("PUNCH", "TCF1"),
    "5468": ("PUNCH.EV", "TCF1"),
    "5473": ("HARRIER.EV", "TCF2"),
    "5466": ("SAFARI", "TCF2"),
    "5479": ("SAFARI", "TCF2"),
    "5605": ("SAFARI", "TCF2"),
    "5465": ("HARRIER", "TCF2"),
    "5478": ("HARRIER", "TCF2"),
    "5604": ("HARRIER", "TCF2"),
    "5483": ("SAFARI.EV", "TCF2"),
}


ENGINE_MASTER = [
    {"LINE": "TCF1", "Engine Part No": "54850000PTP001", "Model": "Punch MT SA", "TA Code": "3302"},
    {"LINE": "TCF1", "Engine Part No": "54850000PTP002", "Model": "Punch AMT SA", "TA Code": "3404"},
    {"LINE": "TCF1", "Engine Part No": "54970000PTP002", "Model": "Punch TC MCE", "TA Code": "7349"},
    {"LINE": "TCF1", "Engine Part No": "54970000PTP003", "Model": "Punch MCE MT", "TA Code": "3641"},
    {"LINE": "TCF1", "Engine Part No": "54970000PTP004", "Model": "Punch MCE AMT", "TA Code": "3406"},
    {"LINE": "TCF1", "Engine Part No": "54970000PTP005", "Model": "Punch MCE CNG MT", "TA Code": "3627"},
    {"LINE": "TCF1", "Engine Part No": "54970000PTP031", "Model": "Punch MCE CNG AMT", "TA Code": "3403"},
    {"LINE": "TCF1", "Engine Part No": "546816111212", "Model": "Nova", "TA Code": "5468"},
    {"LINE": "TCF2", "Engine Part No": "572900000118", "Model": "Harrier / Safari Diesel AT", "TA Code": ""},
    {"LINE": "TCF2", "Engine Part No": "572900000120", "Model": "Harrier / Safari Diesel MT", "TA Code": ""},
    {"LINE": "TCF2", "Engine Part No": "54780000PTP001", "Model": "Harrier / Safari Petrol TGDI MT", "TA Code": ""},
    {"LINE": "TCF2", "Engine Part No": "54780000PTP002", "Model": "Harrier / Safari Petrol TGDI AT", "TA Code": ""},
    {"LINE": "TCF2", "Engine Part No": "547380400103", "Model": "Harrier EV", "TA Code": "5473"},
]

FLOAT_COLS = [
    "TOTAL_FLOAT",
    "PBS_FLOAT",
    "PBS_TO_POLISHING",
    "POLISHING_TO_TOPCOAT",
    "TOPCOAT_TO_WETSANDING_ROOFBLACK",
    "TOPCOAT_TO_WETSANDING_FRESH",
    "WETSANDING_TO_SEALANT",
    "TOTAL_UPTO_SEALANT",
    "PT_ENTRY_TO_SEALENT",
    "BIW_LIFTING_TO_PT",
    "PT_BYPASS",
]

PF_HEADERS = [
    "MARKET", "PRODUCT_FAMILY", "SALES_DESCRIPTION", "SHORT_VC", "PACK",
    "TOTAL_FLOAT", "PBS_FLOAT", "PBS_TO_POLISHING", "POLISHING_TO_TOPCOAT",
    "TOPCOAT_TO_WETSANDING_ROOFBLACK", "TOPCOAT_TO_WETSANDING_FRESH",
    "WETSANDING_TO_SEALANT", "TOTAL_UPTO_SEALANT", "PT_ENTRY_TO_SEALENT",
    "BIW_LIFTING_TO_PT", "PT_BYPASS", "WIRING_PART_NUMBER", "COCKPIT",
    "ENGINE", "FOR_MODEL_FLOAT",
]


# ═══════════════════════════════════════════════════════════════
# 2. CSS THEME
# ═══════════════════════════════════════════════════════════════

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

/* Global Font Override & Soft Background */
html, body, [class*="css"], .stMarkdown, p, span, h1, h2, h3, h4, h5, h6, button, select, input {
    font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif !important;
}

.stApp {
    animation: fadeIn 0.4s ease-out;
}

@keyframes fadeIn {
    from { opacity: 0; transform: translateY(4px); }
    to { opacity: 1; transform: translateY(0); }
}

/* Hide default streamlit header and footer */
header[data-testid="stHeader"] { display: none !important; }
footer { display: none !important; }

/* Dashboard Header Styling */
.dash-header {
    background: linear-gradient(135deg, #0b1329 0%, #1e293b 60%, #1d4ed8 100%) !important;
    box-shadow: 0 10px 30px -10px rgba(0, 0, 0, 0.3) !important;
    border: 1px solid rgba(255, 255, 255, 0.08) !important;
    padding: 24px 32px !important;
    border-radius: 20px !important;
    margin-bottom: 24px !important;
    transition: all 0.3s ease !important;
}
.dash-header:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 15px 35px -8px rgba(29, 78, 216, 0.2) !important;
}
.dash-header h1 {
    font-size: 2rem !important;
    font-weight: 800 !important;
    letter-spacing: -0.025em !important;
    color: #ffffff !important;
    margin: 0 0 6px 0 !important;
}
.dash-header p {
    color: #94a3b8 !important;
    margin: 0 !important;
    font-size: 0.95rem !important;
    font-weight: 500 !important;
}
.dash-header p b {
    color: #60a5fa !important;
    font-weight: 700;
}

/* Custom Live Status Indicators */
.status-container {
    display: flex;
    gap: 10px;
    margin-top: 14px;
    flex-wrap: wrap;
}
.status-badge {
    background: rgba(255, 255, 255, 0.08) !important;
    border: 1px solid rgba(255, 255, 255, 0.12) !important;
    padding: 6px 14px !important;
    border-radius: 50px !important;
    font-size: 0.8rem !important;
    font-weight: 600 !important;
    color: #ffffff !important;
    display: inline-flex !important;
    align-items: center;
    gap: 6px;
}
.status-badge-green {
    background: rgba(16, 185, 129, 0.12) !important;
    border: 1px solid rgba(16, 185, 129, 0.25) !important;
    color: #34d399 !important;
}
.status-badge-blue {
    background: rgba(59, 130, 246, 0.12) !important;
    border: 1px solid rgba(59, 130, 246, 0.25) !important;
    color: #60a5fa !important;
}
.status-badge-amber {
    background: rgba(245, 158, 11, 0.12) !important;
    border: 1px solid rgba(245, 158, 11, 0.25) !important;
    color: #fbbf24 !important;
}

/* Section Title Styling */
.section-title {
    font-size: 1.25rem !important;
    font-weight: 700 !important;
    color: inherit !important;
    margin: 28px 0 14px 0 !important;
    padding-left: 12px !important;
    border-left: 4px solid #2563eb !important;
    border-bottom: none !important;
    padding-bottom: 0 !important;
}

/* Glassmorphic Metric Cards */
[data-testid="stMetric"] {
    background: var(--secondary-background-color, #ffffff) !important;
    border: 1px solid var(--border-color, #e2e8f0) !important;
    padding: 16px 20px !important;
    border-radius: 16px !important;
    box-shadow: 0 4px 15px -3px rgba(0, 0, 0, 0.04) !important;
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
}
[data-testid="stMetric"]:hover {
    transform: translateY(-3px) !important;
    box-shadow: 0 15px 25px -5px rgba(0, 0, 0, 0.06), 0 8px 8px -5px rgba(0, 0, 0, 0.03) !important;
    border-color: #2563eb !important;
}
[data-testid="stMetricLabel"] {
    font-size: 0.78rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
    font-weight: 700 !important;
    color: #64748b !important;
}
[data-testid="stMetricValue"] {
    font-size: 1.8rem !important;
    font-weight: 800 !important;
    letter-spacing: -0.03em !important;
    color: inherit !important;
}

/* Sidebar Custom Styling */
section[data-testid="stSidebar"] {
    background-color: #0b1329 !important;
    border-right: 1px solid #1e293b !important;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p, 
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 {
    color: #f8fafc !important;
}

/* Premium Button Styling */
div.stButton > button {
    border-radius: 12px !important;
    padding: 10px 24px !important;
    font-weight: 700 !important;
    letter-spacing: -0.01em !important;
    border: 1px solid #e2e8f0 !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02) !important;
    transition: all 0.2s ease-in-out !important;
}
div.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 8px 12px -3px rgba(37, 99, 235, 0.12) !important;
    border-color: #2563eb !important;
    color: #2563eb !important;
}
div.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%) !important;
    color: #ffffff !important;
    border: none !important;
}
div.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #1d4ed8 0%, #1e40af 100%) !important;
    color: #ffffff !important;
    box-shadow: 0 8px 16px -3px rgba(37, 99, 235, 0.25) !important;
}

/* Premium Tabs Styling */
button[data-baseweb="tab"] {
    font-size: 0.98rem !important;
    font-weight: 700 !important;
    color: #64748b !important;
    border-radius: 8px 8px 0 0 !important;
    padding: 12px 20px !important;
    border: none !important;
    transition: all 0.2s ease !important;
}
button[data-baseweb="tab"]:hover {
    color: #0f172a !important;
    background-color: #f1f5f9 !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #2563eb !important;
    background-color: rgba(37, 99, 235, 0.05) !important;
    border-bottom: 3px solid #2563eb !important;
}

/* Custom Card Container Grid for File Status */
.file-status-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
}
.file-status-card {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 16px !important;
    padding: 18px !important;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.02) !important;
    transition: all 0.22s ease !important;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    height: 100%;
}
.file-status-card:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 12px 20px -8px rgba(37, 99, 235, 0.08) !important;
    border-color: #2563eb !important;
}
.card-header-line {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 12px;
}
.card-title {
    font-weight: 700 !important;
    color: #1e293b !important;
    font-size: 0.98rem !important;
}
.card-meta-line {
    margin-top: 8px;
    font-size: 0.8rem !important;
    color: #64748b !important;
    line-height: 1.4 !important;
    border-top: 1px solid #f1f5f9;
    padding-top: 8px;
}

/* Custom Status Pills */
.pill-badge {
    border-radius: 50px !important;
    padding: 3px 10px !important;
    font-weight: 700 !important;
    font-size: 0.72rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.03em !important;
    display: inline-flex !important;
    align-items: center;
    gap: 4px;
}
.pill-badge-success {
    background-color: #dcfce7 !important;
    color: #15803d !important;
    border: 1px solid #bbf7d0 !important;
}
.pill-badge-warning {
    background-color: #fef3c7 !important;
    color: #b45309 !important;
    border: 1px solid #fde68a !important;
}
.pill-badge-info {
    background-color: #dbeafe !important;
    color: #1d4ed8 !important;
    border: 1px solid #bfdbfe !important;
}
.pill-badge-danger {
    background-color: #fee2e2 !important;
    color: #b91c1c !important;
    border: 1px solid #fecaca !important;
}

/* Table Header Styles override */
.stDataFrame th {
    font-size: 0.8rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.02em !important;
}

/* Responsive adjustments for mobile viewports */
@media (max-width: 768px) {
    .dash-header {
        padding: 18px 24px !important;
        border-radius: 16px !important;
        margin-bottom: 20px !important;
    }
    .dash-header h1 {
        font-size: 1.5rem !important;
    }
    .dash-header p {
        font-size: 0.88rem !important;
    }
    .section-title {
        font-size: 1.1rem !important;
        margin: 20px 0 10px 0 !important;
    }
}

/* Model Wise Float Table Styling */
.model-wise-table {
    width: 100% !important;
    border-collapse: collapse !important;
    font-size: 0.82rem !important;
    font-family: 'Plus Jakarta Sans', -apple-system, sans-serif !important;
    color: #1e293b !important;
    background-color: #ffffff !important;
}
.model-wise-table th {
    background-color: #334155 !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    padding: 8px 6px !important;
    border: 1px solid #cbd5e1 !important;
    text-align: center !important;
    white-space: normal !important;
    word-wrap: break-word !important;
    vertical-align: middle !important;
    line-height: 1.2 !important;
}
.model-wise-table td {
    padding: 6px 4px !important;
    border: 1px solid #cbd5e1 !important;
    text-align: center !important;
    vertical-align: middle !important;
}
.model-wise-table tr:hover {
    background-color: #f8fafc !important;
}
.model-wise-table tr.grand-total {
    background-color: #dcfce7 !important;
    color: #0f5132 !important;
    font-weight: 800 !important;
}
.model-wise-table tr.line-total {
    background-color: #e0e7ff !important;
    color: #1e3a8a !important;
    font-weight: 700 !important;
}
</style>
"""



# ═══════════════════════════════════════════════════════════════
# 3. UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def _safe(ft: str) -> str:
    """Convert file type name to a safe filename slug."""
    return ft.replace(" ", "_").lower()


def strip_df(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from all string columns and headers."""
    df.columns = [str(c).strip() for c in df.columns]
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    return df


def save_upload(uploaded_file, file_type: str) -> Path:
    """Persist an uploaded file in the local data directory."""
    DATA_DIR.mkdir(exist_ok=True)
    ext = Path(uploaded_file.name).suffix
    target = DATA_DIR / f"{_safe(file_type)}{ext}"
    with open(target, "wb") as f:
        f.write(uploaded_file.getbuffer())
    # Save upload timestamp metadata
    _save_file_meta(file_type, uploaded_file.name)
    return target


def _get_meta_path() -> Path:
    """Return path to file metadata JSON."""
    return DATA_DIR / "file_meta.json"


def _load_file_meta() -> dict:
    """Load file metadata (upload times, original names)."""
    p = _get_meta_path()
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_file_meta(file_type: str, original_name: str):
    """Save upload timestamp and original filename for a file type."""
    DATA_DIR.mkdir(exist_ok=True)
    meta = _load_file_meta()
    meta[file_type] = {
        "uploaded_at": get_ist_now().strftime("%Y-%m-%d %H:%M:%S"),
        "original_name": original_name,
    }
    with open(_get_meta_path(), "w") as f:
        json.dump(meta, f, indent=2)


# ═══════════════════════════════════════════════════════════════
# DAILY 6:30 AM RESET LOGIC
# ═══════════════════════════════════════════════════════════════

def _get_last_reset() -> str | None:
    """Return last recorded reset date string (YYYY-MM-DD), or None."""
    if RESET_STATE_FILE.exists():
        try:
            with open(RESET_STATE_FILE) as f:
                data = json.load(f)
                return data.get("last_reset_date")
        except Exception:
            pass
    return None


def _save_last_reset(date_str: str):
    """Persist the date of the last reset."""
    DATA_DIR.mkdir(exist_ok=True)
    with open(RESET_STATE_FILE, "w") as f:
        json.dump({"last_reset_date": date_str, "reset_time": get_ist_now().strftime("%Y-%m-%d %H:%M:%S")}, f, indent=2)


def check_and_perform_daily_reset():
    """
    If current local time is past 06:30 and we have not yet reset today,
    delete all uploaded data files and the engine stock JSON.
    This ensures a clean slate for each morning shift.
    """
    now = get_ist_now()
    today_str = now.strftime("%Y-%m-%d")
    last_reset = _get_last_reset()

    # Only reset if it's past 06:30 today AND we haven't already reset today
    reset_time_today = now.replace(hour=RESET_HOUR, minute=RESET_MINUTE, second=0, microsecond=0)
    if now >= reset_time_today and last_reset != today_str:
        _perform_reset()
        _save_last_reset(today_str)
        return True  # Reset was performed
    return False  # No reset needed


def _perform_reset():
    """Delete all uploaded data files, engine stock JSON, and file metadata."""
    DATA_DIR.mkdir(exist_ok=True)
    deleted = []
    # Delete all uploaded excel files
    for ext in ["*.xlsx", "*.xlsb", "*.xls"]:
        for f in DATA_DIR.glob(ext):
            try:
                f.unlink()
                deleted.append(f.name)
            except Exception:
                pass
    # Reset engine stock
    if ENGINE_FILE.exists():
        try:
            ENGINE_FILE.unlink()
            deleted.append(ENGINE_FILE.name)
        except Exception:
            pass
    # Reset BOM cache
    if BOM_FILE.exists():
        try:
            BOM_FILE.unlink()
            deleted.append(BOM_FILE.name)
        except Exception:
            pass
    # Reset persistent json caches
    for f in DATA_DIR.glob("cache_*.json"):
        try:
            f.unlink()
        except Exception:
            pass
    # Reset file metadata
    meta_path = _get_meta_path()
    if meta_path.exists():
        try:
            meta_path.unlink()
        except Exception:
            pass
    # Clear session state uploader keys and signatures to prevent stale uploads
    keys_to_clear = [k for k in st.session_state.keys() if k.startswith("last_processed_") or k.startswith("tab6_upload_")]
    for k in keys_to_clear:
        try:
            del st.session_state[k]
        except Exception:
            pass
    # Clear Streamlit cache so fresh data is loaded
    st.cache_data.clear()


# ═══════════════════════════════════════════════════════════════
# BOM MASTER DATA
# ═══════════════════════════════════════════════════════════════

def load_bom_master() -> pd.DataFrame:
    """
    Load BOM (Bill of Materials) master data from Bom details.xlsx.
    Searches recursively in the project root and subdirectories (like Project).
    Also supports fallback to 'Part Number Master' sheet of consolidated TCF workbook.
    Saves a JSON cache in data/ for quick reload.
    Returns a DataFrame with columns: SHORT_VC, FRONT_WIRING, COCKPIT, ENGINE.
    """
    # Find BOM file in root or Project recursively
    bom_path = Path("Bom details.xlsx")
    if not bom_path.exists():
        # Search recursively
        root_dir = Path("d:/TML PPC Dashboard")
        if (root_dir / "Bom details.xlsx").exists():
            bom_path = root_dir / "Bom details.xlsx"
        elif (root_dir / "Project" / "Project files" / "Bom details.xlsx").exists():
            bom_path = root_dir / "Project" / "Project files" / "Bom details.xlsx"
        else:
            matches = [f for f in root_dir.rglob("Bom details.xlsx") if "data" not in f.parts and "test_runs" not in f.parts]
            if matches:
                bom_path = max(matches, key=os.path.getmtime)

    # Try loading from cached JSON if it exists and matches modification time of source Excel
    if bom_path.exists():
        try:
            mtime = os.path.getmtime(bom_path)
            if BOM_FILE.exists():
                with open(BOM_FILE) as f:
                    cached = json.load(f)
                if cached.get("source") == str(bom_path) and cached.get("source_mtime") == mtime:
                    df = pd.DataFrame(cached.get("data", []))
                    if not df.empty:
                        for col in ["SHORT_VC", "FRONT_WIRING", "COCKPIT", "ENGINE"]:
                            if col in df.columns:
                                df[col] = df[col].astype(str).str.strip().replace({"nan": np.nan, "None": np.nan, "": np.nan})
                        return df
        except Exception:
            pass

    # Try loading from BOM source file directly
    if bom_path.exists():
        try:
            df = pd.read_excel(str(bom_path))
            df.columns = [str(c).strip() for c in df.columns]
            # Normalise column names
            rename = {}
            for c in df.columns:
                cu = c.upper().strip()
                if "SHORT" in cu and "VEHICLE" in cu:
                    rename[c] = "SHORT_VC"
                elif "FRONT" in cu and "WIRING" in cu:
                    rename[c] = "FRONT_WIRING"
                elif "COCKPIT" in cu:
                    rename[c] = "COCKPIT"
                elif "ENGINE" in cu:
                    rename[c] = "ENGINE"
            df.rename(columns=rename, inplace=True)
            needed = ["SHORT_VC", "FRONT_WIRING", "COCKPIT", "ENGINE"]
            for n in needed:
                if n not in df.columns:
                    df[n] = np.nan
            df = df[needed].copy()
            # Clean strings
            for col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace({"nan": np.nan, "None": np.nan, "": np.nan})
            df = df.dropna(subset=["SHORT_VC"])
            # Save to data dir as JSON for record-keeping
            DATA_DIR.mkdir(exist_ok=True)
            records = df.fillna("").to_dict("records")
            with open(BOM_FILE, "w") as f:
                json.dump({
                    "source": str(bom_path),
                    "source_mtime": os.path.getmtime(bom_path),
                    "loaded_at": get_ist_now().strftime("%Y-%m-%d %H:%M:%S"),
                    "row_count": len(df),
                    "data": records
                }, f, indent=2)
            return df
        except Exception as e:
            pass  # Fall through to consolidated or cache

    # Fallback to consolidated TCF workbook Part Number Master sheet
    root_dir = Path("d:/TML PPC Dashboard")
    matches = [f for f in root_dir.rglob("*.xlsx") if "TCF VIN" in f.name.upper() and "FLOAT" in f.name.upper() and "MAPPING" in f.name.upper()]
    if matches:
        con_path = max(matches, key=os.path.getmtime)
        try:
            df = pd.read_excel(str(con_path), sheet_name="Part Number Master")
            df.columns = [str(c).strip() for c in df.columns]
            rename = {}
            for c in df.columns:
                cu = c.upper().strip()
                if "SHORT" in cu and "VEHICLE" in cu:
                    rename[c] = "SHORT_VC"
                elif "FRONT" in cu and "WIRING" in cu:
                    rename[c] = "FRONT_WIRING"
                elif "COCKPIT" in cu:
                    rename[c] = "COCKPIT"
                elif "ENGINE" in cu:
                    rename[c] = "ENGINE"
            df.rename(columns=rename, inplace=True)
            needed = ["SHORT_VC", "FRONT_WIRING", "COCKPIT", "ENGINE"]
            for n in needed:
                if n not in df.columns:
                    df[n] = np.nan
            df = df[needed].copy()
            for col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace({"nan": np.nan, "None": np.nan, "": np.nan})
            df = df.dropna(subset=["SHORT_VC"])
            DATA_DIR.mkdir(exist_ok=True)
            records = df.fillna("").to_dict("records")
            with open(BOM_FILE, "w") as f:
                json.dump({
                    "source": str(con_path),
                    "loaded_at": get_ist_now().strftime("%Y-%m-%d %H:%M:%S"),
                    "row_count": len(df),
                    "data": records
                }, f, indent=2)
            return df
        except Exception:
            pass

    # Fallback: load from cached JSON
    if BOM_FILE.exists():
        try:
            with open(BOM_FILE) as f:
                cached = json.load(f)
            df = pd.DataFrame(cached.get("data", []))
            if not df.empty:
                return df
        except Exception:
            pass

    return pd.DataFrame(columns=["SHORT_VC", "FRONT_WIRING", "COCKPIT", "ENGINE"])


def get_bom_info() -> dict:
    """Return metadata about the loaded BOM file."""
    if BOM_FILE.exists():
        try:
            with open(BOM_FILE) as f:
                cached = json.load(f)
            return {
                "loaded_at": cached.get("loaded_at", "Unknown"),
                "row_count": cached.get("row_count", 0),
                "source": Path(cached.get("source", "Unknown")).name,
            }
        except Exception:
            pass
    # Try finding the source file dynamically to show its metadata
    bom_path = Path("Bom details.xlsx")
    if not bom_path.exists():
        root_dir = Path("d:/TML PPC Dashboard")
        matches = [f for f in root_dir.rglob("Bom details.xlsx") if "data" not in f.parts and "test_runs" not in f.parts]
        if matches:
            bom_path = max(matches, key=os.path.getmtime)
    if bom_path.exists():
        mtime = datetime.fromtimestamp(os.path.getmtime(bom_path), tz=timezone.utc).astimezone(timezone(timedelta(hours=5, minutes=30)))
        return {
            "loaded_at": mtime.strftime("%Y-%m-%d %H:%M:%S"),
            "row_count": 0,
            "source": bom_path.name,
        }
    return {}



def _get_last_reset_datetime() -> datetime:
    """Return the datetime of the last reset as a timezone-aware (IST) datetime,
    or a very old datetime (e.g., 1970-01-01) if not set.
    """
    if RESET_STATE_FILE.exists():
        try:
            with open(RESET_STATE_FILE) as f:
                data = json.load(f)
                reset_time_str = data.get("reset_time")
                if reset_time_str:
                    # Parse reset_time_str ("YYYY-MM-DD HH:MM:SS")
                    dt = datetime.strptime(reset_time_str, "%Y-%m-%d %H:%M:%S")
                    # Localize to IST (UTC+5:30)
                    return dt.replace(tzinfo=timezone(timedelta(hours=5, minutes=30)))
        except Exception:
            pass
    # Default to epoch in IST
    return datetime(1970, 1, 1, tzinfo=timezone(timedelta(hours=5, minutes=30)))


@st.cache_data
def _scan_workspace_files(last_reset_dt_str: str) -> list:
    """Scan workspace for Excel files, ignoring files modified before last_reset_dt.
    Uses a fast directory walk with pruning of unwanted directories.
    """
    last_reset_dt = datetime.strptime(last_reset_dt_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone(timedelta(hours=5, minutes=30))) if last_reset_dt_str else datetime(1970, 1, 1, tzinfo=timezone(timedelta(hours=5, minutes=30)))
    
    root_dir = Path("d:/TML PPC Dashboard")
    all_files = []
    if root_dir.exists():
        for root, dirs, files in os.walk(root_dir):
            # Prune directories in place to prevent os.walk from entering them
            dirs[:] = [d for d in dirs if d not in [".git", "data", "test_runs", ".agents", "__pycache__", ".pytest_cache", ".streamlit"]]
            for file in files:
                if file.lower().endswith((".xls", ".xlsx", ".xlsb")):
                    f = Path(root) / file
                    try:
                        mtime = datetime.fromtimestamp(os.path.getmtime(f), tz=timezone.utc).astimezone(timezone(timedelta(hours=5, minutes=30)))
                        if mtime >= last_reset_dt:
                            all_files.append(f)
                    except Exception:
                        continue
    return all_files


def load_with_persistent_cache(cache_key: str, src_path: str, parse_func, *args, **kwargs) -> pd.DataFrame:
    """Load parsed data from a persistent JSON cache if it exists and is up to date,
    otherwise parse and save to JSON cache.
    """
    cache_meta_file = DATA_DIR / f"cache_{cache_key}_meta.json"
    cache_data_file = DATA_DIR / f"cache_{cache_key}.json"
    
    try:
        current_mtime = os.path.getmtime(src_path)
    except Exception:
        current_mtime = 0.0

    if cache_meta_file.exists() and cache_data_file.exists():
        try:
            with open(cache_meta_file, "r") as f:
                meta = json.load(f)
            if meta.get("file_mtime") == current_mtime:
                df = pd.read_json(cache_data_file, orient="split")
                return df
        except Exception:
            pass

    # Cache miss
    df = parse_func(*args, **kwargs)
    
    # Save to persistent cache
    try:
        DATA_DIR.mkdir(exist_ok=True)
        with open(cache_meta_file, "w") as f:
            json.dump({"file_mtime": current_mtime}, f)
        df.to_json(cache_data_file, orient="split")
    except Exception:
        pass
        
    return df


def get_source(file_type: str):
    """Return (source_type, path) — 'uploaded' or 'scanned' or (None, None).
    Recursively searches project directory (including Project folder) and
    falls back to a consolidated workbook sheet if individual files aren't newer/found.
    """
    # Check for manually uploaded file first
    for ext in [".xlsb", ".xlsx", ".xls"]:
        up = DATA_DIR / f"{_safe(file_type)}{ext}"
        if up.exists():
            return "uploaded", up

    last_reset_dt = _get_last_reset_datetime()
    last_reset_dt_str = last_reset_dt.strftime("%Y-%m-%d %H:%M:%S")

    # Get scanned files from cached workspace scanner
    all_files = _scan_workspace_files(last_reset_dt_str)

    # Keywords/patterns to match individual files by file_type
    patterns = {
        "Paint Float": [r"PPC_Float_Report", r"Paint float report", r"Paint Float"],
        "TCF1 DPT Plan": [r"DPT_PLAN-VIN_GENERATION_REPORT", r"TCF 1 DPT plan", r"TCF1_DPT", r"TCF1 DPT"],
        "TCF2 DPT Plan": [r"TCF2_DPT-PLAN", r"TCF2_DPT-PLAN_VIN_GENERATION_REPORT", r"TCF2 DPT", r"TCF 2 DPT"],
        "TCF1 Cockpit": [r"TCF1_Cockpit", r"TCF 1 COCKPIT", r"TCF1 Cockpit"],
        "TCF2 Cockpit": [r"Harrier safari cockpit", r"TCF 2 cockpit", r"TCF2 Cockpit"],
        "Nova Cockpit": [r"Nova_Cockpit", r"Nova"],
        "TCF1 Wiring File": [r"Wiring Harness report NEW", r"TCF 1 Wiring", r"TCF 1 Wiring Harness", r"TCF1 Wiring"],
        "TCF2 Wiring File": [r"TCF2_Wiring", r"TCF-2 Wiring", r"TCF-2 Wiring Harness", r"TCF2 Wiring"],
        "Engine Manual Data": [r"Engine manual data", r"Engine manual", r"Engine stock", r"Engine_manual_data"]
    }

    if file_type in patterns:
        individual_candidates = []
        consolidated_candidates = []
        for f in all_files:
            name = f.name.upper()
            is_consolidated = ("TCF VIN" in name and "FLOAT" in name and "MAPPING" in name)
            if is_consolidated:
                consolidated_candidates.append((f, os.path.getmtime(f)))
                continue
            
            matched = False
            for pat in patterns[file_type]:
                if re.search(pat.upper(), name):
                    matched = True
                    break
            if matched:
                individual_candidates.append((f, os.path.getmtime(f)))

        # Prefer individual files if found, otherwise fall back to consolidated
        if individual_candidates:
            individual_candidates.sort(key=lambda x: x[1], reverse=True)
            return "scanned", individual_candidates[0][0]
        elif consolidated_candidates:
            consolidated_candidates.sort(key=lambda x: x[1], reverse=True)
            return "scanned", consolidated_candidates[0][0]

    return None, None


def _is_html_xls(path) -> bool:
    """Detect if a file is an HTML table disguised with an .xls/.xlsx extension."""
    try:
        with open(path, "rb") as f:
            header = f.read(16)
        # HTML files start with '<' or BOM+<
        return header.lstrip(b"\xef\xbb\xbf").lstrip().startswith(b"<")
    except Exception:
        return False


def _read_html_table_bs4(path) -> pd.DataFrame:
    """
    Parse an HTML file using BeautifulSoup.
    Handles invalid colspan=0 / rowspan=0 attributes that break pd.read_html.
    Returns a DataFrame with the first table's data.
    """
    from bs4 import BeautifulSoup
    with open(path, "rb") as f:
        raw = f.read().decode("utf-8", errors="replace")
    soup = BeautifulSoup(raw, "lxml")
    table = soup.find("table")
    if table is None:
        return pd.DataFrame()
    rows = table.find_all("tr")
    all_rows = []
    for r in rows:
        cells = [td.get_text(strip=True) for td in r.find_all(["td", "th"])]
        all_rows.append(cells)
    if not all_rows:
        return pd.DataFrame()
    # Normalize row lengths
    max_cols = max(len(r) for r in all_rows)
    for r in all_rows:
        while len(r) < max_cols:
            r.append("")
    return pd.DataFrame(all_rows)


def read_sheet(src_type, src_path, sheet_name, **kw):
    """
    Read a sheet from an uploaded/scanned file.

    Handles three formats transparently:
      - .xlsb        → pyxlsb engine
      - .xls/.xlsx   → openpyxl / xlrd
      - HTML-as-XLS  → BeautifulSoup fallback (web-exported reports)
    """
    src_path_str = str(src_path)

    # HTML-disguised file? (web report saved as .xls)
    if _is_html_xls(src_path_str):
        return _read_html_table_bs4(src_path_str)

    engine = "pyxlsb" if src_path_str.lower().endswith(".xlsb") else None
    if src_type in ["uploaded", "scanned"]:
        try:
            return pd.read_excel(src_path, sheet_name=sheet_name, engine=engine, **kw)
        except Exception:
            return pd.read_excel(src_path, engine=engine, **kw)
    return pd.read_excel(src_path, sheet_name=sheet_name, engine=engine, **kw)


def to_excel(df_or_styler, sheet="Summary", table_type="generic") -> bytes:
    """Convert a DataFrame or Styler to downloadable xlsx bytes, adding openpyxl formatting."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    is_styler = hasattr(df_or_styler, "to_excel") and hasattr(df_or_styler, "data")
    df = df_or_styler.data.copy() if is_styler else df_or_styler.copy()

    # Drop Today VIN from wiring and cockpit outputs to match the user shortage report image
    if table_type in ["wiring", "cockpit"]:
        df = df.drop(columns=["Today VIN"], errors="ignore")
        if table_type == "cockpit":
            df = df.rename(columns={
                "Model/Line": "Model",
                "Cabs FloatUPTO SEALANT": "Cabs Float UPTO SEALANT",
                "Shortage for TOTAL FLOAT": "Shortage TOTAL FLOAT"
            })

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if table_type == "engine":
            # Custom Engine & Battery multi-level formatting
            cols_ordered = ["LINE", "Engine Part No", "Model", "TA Code", "Clearance After 6:30AM", "Today VIN", "Bal", "PBS FLOAT", "Float UPTO SEALANT", "TOTAL FLOAT", "With respect to PBS FLOAT", "With respect to Sealant FLOAT", "With respect to Total FLOAT"]
            df = df[[c for c in cols_ordered if c in df.columns]]
            df.to_excel(w, index=False, startrow=2, header=False, sheet_name=sheet)
            ws = w.sheets[sheet]
            
            peach = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            pink = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
            cyan = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
            yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            headers = [
                ("LINE", "LINE", peach),
                ("Manual data", "Engine Part No", peach),
                ("Manual data", "Model", peach),
                ("Manual data", "TA Code", peach),
                ("Manual data", "Clearance After 6:30AM", green),
                ("Manual data", "Today VIN", peach),
                ("Manual data", "Bal", pink),
                ("Paint Float", "PBS FLOAT", peach),
                ("Paint Float", "Float UPTO SEALANT", peach),
                ("Paint Float", "TOTAL FLOAT", peach),
                ("Engine & Battery requirement", "With respect to PBS FLOAT", peach),
                ("Engine & Battery requirement", "With respect to Sealant FLOAT", peach),
                ("Engine & Battery requirement", "With respect to Total FLOAT", peach),
            ]
            
            for i, (top, bottom, fill) in enumerate(headers):
                col_idx = i + 1
                c1 = ws.cell(row=1, column=col_idx, value=top)
                c2 = ws.cell(row=2, column=col_idx, value=bottom)
                c1.border = border; c1.alignment = center_align; c1.font = Font(bold=True)
                c2.border = border; c2.alignment = center_align; c2.font = Font(bold=True)
                c1.fill = fill; c2.fill = fill
                
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
            ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
            ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            
            for i, col_name in enumerate(df.columns):
                col_idx = i + 1
                col_letter = get_column_letter(col_idx)
                max_len = len(str(col_name))
                for row in range(3, len(df) + 3):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    val = str(cell.value or "")
                    if len(val) > max_len: max_len = len(val)
                    
                    if col_name == "Clearance After 6:30AM": cell.fill = green
                    elif col_name == "Bal": cell.fill = pink
                    elif "With respect to" in col_name and isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.fill = red_fill
                        cell.font = Font(color="8B0000", bold=True)
                        
                    model_val = str(ws.cell(row=row, column=3).value or "")
                    if "Total" in model_val: cell.fill = cyan
                    elif model_val in ["TCF1", "TCF2"]: cell.fill = yellow
                        
                ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

            start_tcf1 = end_tcf1 = start_tcf2 = end_tcf2 = None
            for row in range(3, len(df) + 3):
                line_val = ws.cell(row=row, column=1).value
                if line_val == "TCF1":
                    if not start_tcf1: start_tcf1 = row
                    end_tcf1 = row
                elif line_val == "TCF2":
                    if not start_tcf2: start_tcf2 = row
                    end_tcf2 = row
            if start_tcf1 and end_tcf1 > start_tcf1: ws.merge_cells(start_row=start_tcf1, start_column=1, end_row=end_tcf1, end_column=1)
            if start_tcf2 and end_tcf2 > start_tcf2: ws.merge_cells(start_row=start_tcf2, start_column=1, end_row=end_tcf2, end_column=1)

        else:
            # Generic / Wiring formatting
            df.to_excel(w, index=False, sheet_name=sheet)
            ws = w.sheets[sheet]
            peach_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            generic_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            red_font = Font(color="C00000", bold=True)
            
            thin_side = Side(border_style="thin", color="000000")
            double_side = Side(border_style="double", color="000000")

            for i, col_name in enumerate(df.columns):
                col_idx = i + 1
                col_letter = get_column_letter(col_idx)
                
                # Header formatting
                header_cell = ws.cell(row=1, column=col_idx)
                header_cell.border = Border(
                    top=thin_side, bottom=thin_side, left=thin_side,
                    right=double_side if i in [2, 5] else thin_side
                )
                header_cell.alignment = center_align
                header_cell.font = Font(bold=True)
                
                if table_type in ["wiring", "cockpit"]:
                    if i < 3: header_cell.fill = peach_fill
                    else: header_cell.fill = blue_fill
                else:
                    header_cell.fill = generic_fill
                    
                max_len = len(str(col_name))
                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = Border(
                        top=thin_side, bottom=thin_side, left=thin_side,
                        right=double_side if i in [2, 5] else thin_side
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    val = str(cell.value or "")
                    if len(val) > max_len: max_len = len(val)
                    
                    # Highlight negative shortages
                    if table_type in ["wiring", "cockpit", "generic"] and ("Shortage" in str(col_name) or "With respect to" in str(col_name)):
                        try:
                            num_val = float(cell.value)
                            if num_val < 0:
                                cell.fill = red_fill
                                cell.font = red_font
                        except (ValueError, TypeError):
                            pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

    return buf.getvalue()


def export_df_to_image(df: pd.DataFrame, table_type="wiring") -> bytes:
    """Render the shortage report dataframe to a high-fidelity image matching the Excel style."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from io import BytesIO

    plot_df = df.copy()
    
    # Drop Today VIN from wiring and cockpit outputs if it exists to match the screenshot
    if "Today VIN" in plot_df.columns:
        plot_df = plot_df.drop(columns=["Today VIN"], errors="ignore")
        
    if table_type == "cockpit":
        plot_df = plot_df.rename(columns={
            "Model/Line": "Model",
            "Cabs FloatUPTO SEALANT": "Cabs Float UPTO SEALANT",
            "Shortage for TOTAL FLOAT": "Shortage TOTAL FLOAT"
        })

    # Wrap header text exactly to match layouts
    wrapped_headers = []
    for c in plot_df.columns:
        col_name = str(c)
        if "Part Number" in col_name:
            wrapped_headers.append(col_name.replace(" Part Number", "\nPart Number"))
        elif "Clearance After" in col_name:
            wrapped_headers.append("Clearance\nAfter\n6:30AM")
        elif "Paint TOTAL" in col_name:
            wrapped_headers.append("Paint TOTAL\nFLOAT")
        elif "PBS FLOAT" in col_name:
            wrapped_headers.append("PBS\nFLOAT")
        elif "Cabs Float" in col_name:
            wrapped_headers.append("Cabs Float\nUPTO\nSEALANT")
        elif "Shortage PBS" in col_name:
            wrapped_headers.append("Shortage PBS\nFLOAT")
        elif "Upto Sealant" in col_name:
            wrapped_headers.append("Shortage\nUpto\nSealant")
        elif "TOTAL FLOAT" in col_name:
            if "Shortage for" in col_name:
                wrapped_headers.append("Shortage\nfor TOTAL\nFLOAT")
            else:
                wrapped_headers.append("Shortage\nTOTAL\nFLOAT")
        else:
            wrapped_headers.append(col_name)

    # Shortage columns to format
    shortage_cols = [c for c in plot_df.columns if "Shortage" in str(c) or "With respect to" in str(c)]
    shortage_indices = [plot_df.columns.get_loc(c) for c in shortage_cols]

    num_cols = len(plot_df.columns)
    num_rows = len(plot_df)

    header_height = 0.8
    row_height = 0.45
    fig_width = 1.25 * num_cols
    fig_height = header_height + (num_rows * row_height)

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.axis("off")

    # Construct the table
    tbl = ax.table(
        cellText=plot_df.astype(str).values,
        colLabels=wrapped_headers,
        loc="center",
        bbox=[0, 0, 1, 1]
    )

    total_units = header_height + (num_rows * row_height)
    h_header = header_height / total_units
    h_row = row_height / total_units

    peach_color = "#F8CBAD"
    blue_color = "#BDD7EE"
    pink_color = "#FCE4D6"
    red_text = "#C00000"

    for (r, c), cell in tbl.get_celld().items():
        cell.set_linewidth(0.5)
        cell.set_edgecolor("#000000")
        
        if r == 0:
            cell.set_height(h_header)
            cell.set_y(1.0 - h_header)
            cell.set_facecolor(peach_color if c < 3 else blue_color)
            cell.set_text_props(weight="bold", fontfamily="sans-serif", fontsize=10)
        else:
            cell.set_height(h_row)
            cell.set_y(1.0 - h_header - r * h_row)
            cell.set_text_props(fontfamily="sans-serif", fontsize=10)
            
            # Formatting negative shortage cells
            if c in shortage_indices:
                try:
                    val = float(plot_df.iloc[r - 1, c])
                    if val < 0:
                        cell.set_facecolor(pink_color)
                        cell.get_text().set_color(red_text)
                        cell.set_text_props(weight="bold", fontfamily="sans-serif", fontsize=10)
                except:
                    pass

    # Double vertical lines separation
    # Boundaries: between col 3 (index 2) and col 4 (index 3) at x = 3 / num_cols
    # and between col 6 (index 5) and col 7 (index 6) at x = 6 / num_cols
    x1 = 3.0 / num_cols
    x2 = 6.0 / num_cols
    offset = 0.0018
    ax.axvline(x=x1 - offset, color="#000000", linewidth=0.8, zorder=10)
    ax.axvline(x=x1 + offset, color="#000000", linewidth=0.8, zorder=10)
    ax.axvline(x=x2 - offset, color="#000000", linewidth=0.8, zorder=10)
    ax.axvline(x=x2 + offset, color="#000000", linewidth=0.8, zorder=10)

    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def load_engine_from_excel(path: Path) -> dict:
    """Parse the Engine Manual Data sheet and return a dict of engine_part_no -> clearance."""
    try:
        # Since it might be an HTML disguised as XLS or a normal Excel, use read_sheet wrapper
        df = read_sheet("uploaded", path, 0)
        if df.empty:
            return {}
        
        # Find header row
        h = _find_header_row(df, ["ENGINE PART", "CLEARANCE"])
        if h is None:
            # Let's try finding by checking if any row contains "Engine Part No"
            for i in range(min(15, len(df))):
                row_vals = [str(v).lower().strip() for v in df.iloc[i]]
                if any("engine part" in v for v in row_vals):
                    h = i
                    break
        
        if h is not None:
            # Set columns to header row
            df.columns = [str(c).strip() for c in df.iloc[h]]
            df = df.iloc[h+1:].copy().reset_index(drop=True)
        else:
            # Fallback: assume column 0 is Engine Part No and column 3 is Clearance
            df.columns = ["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM"] + list(df.columns[4:])
        
        # Rename columns to standard names
        rename_map = {}
        for col in df.columns:
            col_upper = str(col).upper()
            if "ENGINE PART" in col_upper:
                rename_map[col] = "Engine Part No"
            elif "CLEARANCE" in col_upper:
                rename_map[col] = "Clearance After 6:30AM"
            elif "TA CODE" in col_upper:
                rename_map[col] = "TA Code"
            elif "MODEL" in col_upper:
                rename_map[col] = "Model"
        df = df.rename(columns=rename_map)
        
        # Ensure correct columns exist
        if "Engine Part No" not in df.columns:
            return {}
            
        clearance_col = "Clearance After 6:30AM" if "Clearance After 6:30AM" in df.columns else df.columns[min(3, len(df.columns)-1)]
        
        result = {}
        for _, row in df.iterrows():
            ep = str(row["Engine Part No"]).strip()
            # A valid engine part number should have length >= 5 and not look like a total
            if ep and len(ep) >= 5 and "TOTAL" not in ep.upper() and "TCF" not in ep.upper():
                clearance = 0
                try:
                    val = row.get(clearance_col, 0)
                    if pd.notna(val):
                        clearance = int(float(val))
                except Exception:
                    pass
                result[ep] = clearance
        return result
    except Exception as e:
        print(f"Error parsing engine excel: {e}")
        return {}


def load_engine_json() -> dict:
    """Load persisted engine stock data, falling back to uploaded/scanned Excel file if newer."""
    data = {}
    json_mtime = 0
    if ENGINE_FILE.exists():
        try:
            json_mtime = os.path.getmtime(ENGINE_FILE)
            with open(ENGINE_FILE) as f:
                data = json.load(f)
        except Exception:
            pass
            
    # Check if there is an uploaded or scanned "Engine Manual Data" sheet
    src_type, src_path = get_source("Engine Manual Data")
    if src_type is not None:
        try:
            excel_mtime = os.path.getmtime(src_path)
            # If Excel is newer than json, or json doesn't exist/is empty, load from Excel
            if excel_mtime > json_mtime or not data:
                excel_data = load_engine_from_excel(src_path)
                for ep, clearance in excel_data.items():
                    if ep not in data:
                        data[ep] = {"model": "", "ta_code": "", "clearance": clearance}
                    else:
                        data[ep]["clearance"] = clearance
                # Save the merged data back to JSON so that it's persisted and the JSON mtime is updated
                save_engine_json(data)
        except Exception:
            pass
                
    return data


def save_engine_json(data: dict):
    """Persist engine stock data to JSON."""
    DATA_DIR.mkdir(exist_ok=True)
    with open(ENGINE_FILE, "w") as f:
        json.dump(data, f, indent=2)


@st.dialog("Engine Manual Entry Sheet", width="large")
def open_manual_entry_popup(editor_df):
    st.markdown("Edit the **Clearance After 6:30AM** values below. Click **Save & Close** to apply changes.")

    # Reorder columns: make TA Code last
    cols_order = ["Engine Part No", "Model", "Clearance After 6:30AM", "TA Code"]
    editor_df = editor_df[cols_order]

    edited = st.data_editor(
        editor_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        disabled=["Engine Part No", "Model", "TA Code"],
        column_config={
            "Engine Part No": st.column_config.TextColumn("Engine Part No", width="medium"),
            "Model": st.column_config.TextColumn("Model", width="large"),
            "Clearance After 6:30AM": st.column_config.NumberColumn("Clearance After 6:30AM", min_value=0, step=1, width="medium"),
            "TA Code": st.column_config.TextColumn("TA Code", width="small"),
        },
        key="engine_editor_modal",
    )

    st.markdown("")
    col_save, _ = st.columns([1, 4])
    if col_save.button("💾 Save & Close", type="primary"):
        # Load existing JSON from disk to merge and preserve other fields (today_vin, etc.)
        current_json = load_engine_json()
        
        # Merge updated values
        for _, row in edited.iterrows():
            ep = str(row.get("Engine Part No", "")).strip()
            if not ep or ep == "nan":
                continue
            clearance_val = int(row.get("Clearance After 6:30AM", 0) or 0)
            if ep in current_json:
                current_json[ep]["clearance"] = clearance_val
            else:
                current_json[ep] = {
                    "model": str(row.get("Model", "")),
                    "ta_code": str(row.get("TA Code", "")),
                    "clearance": clearance_val
                }
                
        save_engine_json(current_json)
        
        # Clear the session state cached dataframe so it reloads fresh data from JSON next time
        if "engine_editor_df" in st.session_state:
            del st.session_state["engine_editor_df"]
            
        st.success("Saved successfully!")
        st.rerun()


def is_valid_vc(val) -> bool:
    """Check if a value looks like a valid Short Vehicle Code."""
    s = str(val).strip()
    return bool(re.match(r"^\d{7,9}[A-Za-z]?$", s))


# ═══════════════════════════════════════════════════════════════
# 4. PARSING FUNCTIONS
# ═══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner="Loading BOM Master…")
def load_bom_master_cached() -> pd.DataFrame:
    """Cached wrapper for BOM master load."""
    return load_bom_master()


def _find_header_row(df_raw, keywords, max_scan=15):
    """Find the row index containing at least one keyword."""
    for i in range(min(max_scan, len(df_raw))):
        vals = [str(v).upper().strip() for v in df_raw.iloc[i] if pd.notna(v)]
        for kw in keywords:
            if any(kw in v for v in vals):
                return i
    return None


@st.cache_data
def parse_paint_float(src_type, src_path, pm_df=None, file_mtime: float = 0.0) -> pd.DataFrame:
    """Parse Paint Float into a clean DataFrame with model/line assignments."""
    df = read_sheet(src_type, src_path, "Paint Float", header=None)

    # Locate the header row (contains MARKET, PRODUCT FAMILY, or SHORT VC)
    h = _find_header_row(df, ["MARKET", "PRODUCT FAMILY", "SHORT VC"])
    if h is None:
        h = 2  # fallback

    data = df.iloc[h + 1 :].copy().reset_index(drop=True)

    # Pad / trim to 16 columns
    while len(data.columns) < 16:
        data[len(data.columns)] = np.nan
    data = data.iloc[:, :16]
    data.columns = PF_HEADERS[:16]
    
    # Initialize extra columns
    data["WIRING_PART_NUMBER"] = ""
    data["COCKPIT"] = ""
    data["ENGINE"] = ""
    data["FOR_MODEL_FLOAT"] = ""

    # Strip strings
    for col in data.columns:
        if data[col].dtype == "object":
            data[col] = data[col].astype(str).str.strip()
            data[col] = data[col].replace({"nan": np.nan, "None": np.nan, "": np.nan, "0": np.nan, "0.0": np.nan})

    # Keep only valid data rows
    data = data[data["SHORT_VC"].notna()]
    data = data[~data.get("PRODUCT_FAMILY", pd.Series(dtype=str)).str.contains("Total", na=False, case=False)]
    data = data[~data["SHORT_VC"].str.contains("Total|Grand|TCF", na=False, case=False)]

    # Merge PM Data
    if pm_df is not None and not pm_df.empty:
        pm_subset = pm_df[["SHORT_VC", "FRONT_WIRING", "COCKPIT", "ENGINE"]].drop_duplicates("SHORT_VC")
        data = data.drop(columns=["WIRING_PART_NUMBER", "COCKPIT", "ENGINE"], errors="ignore")
        data = data.merge(pm_subset, on="SHORT_VC", how="left")
        data.rename(columns={"FRONT_WIRING": "WIRING_PART_NUMBER"}, inplace=True)
    
    # Fill nan after merge
    for c in ["WIRING_PART_NUMBER", "COCKPIT", "ENGINE"]:
        if c not in data: data[c] = ""
        data[c] = data[c].fillna("")

    # Numeric conversions
    for col in FLOAT_COLS:
        data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0).astype(int)

    # Model / Line classification from SHORT_VC column via MODEL_MAP
    def classify_by_vc(vc):
        vc_str = str(vc).strip()[:4]
        for k, (mod, ln) in MODEL_MAP.items():
            if vc_str.startswith(k[:4]):
                return pd.Series([mod, ln])
        return pd.Series(["OTHER", "UNKNOWN"])

    data[["MODEL", "LINE"]] = data["SHORT_VC"].apply(classify_by_vc)
    data = data[data["LINE"] != "UNKNOWN"].reset_index(drop=True)
    return data


@st.cache_data
def parse_dpt_plan(src_type, src_path, sheet_name, line_label, file_mtime: float = 0.0) -> pd.DataFrame:
    """Parse a DPT Plan VIN Generation Report.

    Supports both real .xlsb files and HTML-as-.xls exports from web portals.
    HTML files use invalid colspan/rowspan=0 that breaks pd.read_html — we use
    BeautifulSoup directly instead.

    Returns columns: SHORT_VC, SALES_DESC, DPT_PLAN, DPT_VIN, WIRING, COCKPIT, ENGINE, FOR_MODEL, LINE
    """
    src_path_str = str(src_path)

    # ── HTML-based DPT file (web export) ──
    if _is_html_xls(src_path_str):
        raw_df = _read_html_table_bs4(src_path_str)
        if raw_df.empty:
            return pd.DataFrame(columns=["SHORT_VC","SALES_DESC","DPT_PLAN","DPT_VIN","WIRING","COCKPIT","ENGINE","FOR_MODEL","LINE"])

        # First row is header
        header = [str(v).strip() for v in raw_df.iloc[0]]
        data = raw_df.iloc[1:].copy().reset_index(drop=True)
        data.columns = range(len(data.columns))

        # Map column indices from header
        def _col_idx(keywords):
            for kw in keywords:
                for i, h in enumerate(header):
                    if kw.upper() in h.upper():
                        return i
            return None

        vc_col   = _col_idx(["VC"])
        desc_col = _col_idx(["SALES DESC", "SALES"])
        plan_col = _col_idx(["PLAN", "TCF/-Plan", "TCF2-Plan"])
        vin_col  = _col_idx(["VIN", "TCF/-VIN", "TCF2-VIN"])

        result_rows = []
        for _, row in data.iterrows():
            vc = str(row.iloc[vc_col]).strip() if vc_col is not None and vc_col < len(row) else ""
            desc = str(row.iloc[desc_col]).strip() if desc_col is not None and desc_col < len(row) else ""
            plan_val = row.iloc[plan_col] if plan_col is not None and plan_col < len(row) else 0
            vin_val  = row.iloc[vin_col]  if vin_col is not None  and vin_col  < len(row) else 0

            # Skip header repeats, totals, empty rows
            if not vc or vc.upper() in ("VC", "TOTAL", "GRAND TOTAL", ""):
                continue

            result_rows.append({
                "SHORT_VC":   vc,
                "SALES_DESC": desc,
                "DPT_PLAN":   int(pd.to_numeric(plan_val, errors="coerce") or 0),
                "DPT_VIN":    int(pd.to_numeric(vin_val,  errors="coerce") or 0),
                "WIRING":     "",
                "COCKPIT":    "",
                "ENGINE":     "",
                "FOR_MODEL":  "",
                "LINE":       line_label,
            })

        return pd.DataFrame(result_rows) if result_rows else pd.DataFrame(
            columns=["SHORT_VC","SALES_DESC","DPT_PLAN","DPT_VIN","WIRING","COCKPIT","ENGINE","FOR_MODEL","LINE"])

    # ── Normal Excel / .xlsb file ──
    df = read_sheet(src_type, src_path, sheet_name, header=None)

    h = _find_header_row(df, ["MARKET", "PRODUCTFAMILY", "VC", "SALES"])
    if h is None:
        h = 0

    data = df.iloc[h + 1:].copy().reset_index(drop=True)
    while len(data.columns) < 10:
        data[len(data.columns)] = np.nan
    data = data.iloc[:, :10]
    data.columns = [
        "MARKET", "PRODUCT_FAMILY", "SHORT_VC", "SALES_DESC",
        "DPT_PLAN", "DPT_VIN", "C6", "C7", "C8", "C9"
    ]
    data["WIRING"]    = ""
    data["COCKPIT"]   = ""
    data["ENGINE"]    = ""
    data["FOR_MODEL"] = ""

    for col in data.columns:
        if data[col].dtype == "object":
            data[col] = data[col].astype(str).str.strip()
            data[col] = data[col].replace({"nan": np.nan, "None": np.nan, "": np.nan})

    data["DPT_PLAN"] = pd.to_numeric(data["DPT_PLAN"], errors="coerce").fillna(0).astype(int)
    data["DPT_VIN"]  = pd.to_numeric(data["DPT_VIN"],  errors="coerce").fillna(0).astype(int)
    data = data[data["SHORT_VC"].notna() & (data["SHORT_VC"] != "")].reset_index(drop=True)
    data["LINE"] = line_label
    return data


@st.cache_data
def parse_wiring_tcf1(src_type, src_path, file_mtime: float = 0.0) -> pd.DataFrame:
    """Parse TCF1 Wiring File → part, stock, plan per wiring harness."""
    sheet_name = "TCF1 Wiring File"
    if src_type in ["uploaded", "scanned"]:
        engine = "pyxlsb" if str(src_path).lower().endswith(".xlsb") else None
        try:
            with pd.ExcelFile(src_path, engine=engine) as xls:
                sheets = xls.sheet_names
                if sheet_name not in sheets:
                    for s in sheets:
                        if "Coverage" in s or "coverage" in s:
                            sheet_name = s
                            break
                    if sheet_name not in sheets:
                        sheet_name = sheets[0]
        except Exception:
            pass

    df = read_sheet(src_type, src_path, sheet_name, header=None)

    h = _find_header_row(df, ["VC NUMBER", "TOTAL PLAN"])
    if h is None:
        h = 1  # fallback
    header_row = df.iloc[h]
    data = df.iloc[h + 1 :].copy().reset_index(drop=True)

    col_map = {}
    for j, v in enumerate(header_row):
        if pd.isna(v): continue
        vu = str(v).upper().strip()
        if "VC NUMBER" in vu or "VC NO" in vu: col_map.setdefault("VC", j)
        if "WIRING HARNESS" in vu: col_map.setdefault("WH", j)
        if "TOTAL PLAN" in vu: col_map.setdefault("PLAN", j)
        if "COVERAGE" in vu or "FRESH VIN" in vu or vu == "FRESH VIN STOCK":
            col_map["STOCK"] = j
        elif vu == "STOCK" and "STOCK" not in col_map:
            col_map["STOCK"] = j

    vc_col = col_map.get("VC", 2)
    wh_col = col_map.get("WH", 3)
    plan_col = col_map.get("PLAN", 5)
    stock_col = col_map.get("STOCK", 8)

    result_rows = []
    for _, row in data.iterrows():
        vc_val = str(row.iloc[vc_col]).strip() if vc_col < len(row) and pd.notna(row.iloc[vc_col]) else ""
        wh_val = str(row.iloc[wh_col]).strip() if wh_col < len(row) and pd.notna(row.iloc[wh_col]) else ""
        stock_val = row.iloc[stock_col] if stock_col < len(row) else np.nan
        plan_val = row.iloc[plan_col] if plan_col < len(row) else np.nan

        # Only keep rows with a valid Short VC and wiring part
        if is_valid_vc(vc_val) and len(wh_val) >= 10 and wh_val[0].isdigit():
            result_rows.append({
                "SHORT_VC": vc_val,
                "WIRING_PART": wh_val,
                "STOCK": pd.to_numeric(stock_val, errors="coerce") or 0,
                "PLAN": pd.to_numeric(plan_val, errors="coerce") or 0,
                "LINE": "TCF1",
            })

    return pd.DataFrame(result_rows) if result_rows else pd.DataFrame(
        columns=["SHORT_VC", "WIRING_PART", "STOCK", "PLAN", "LINE"]
    )


@st.cache_data
def parse_wiring_tcf2(src_type, src_path, file_mtime: float = 0.0) -> pd.DataFrame:
    """Parse TCF2 Wiring File → Short VC, stock (needs PM mapping)."""
    df = read_sheet(src_type, src_path, "TCF2 Wiring File", header=None)

    # Structure: Row 0 = title, Row 1 = sub-headers (Plan, STOCK, …)
    # Data from row 2+. col 1 = Short VC, col 8 = STOCK
    h = _find_header_row(df, ["STOCK", "PLAN", "REMARK"])
    if h is None:
        h = 1
    header_row = df.iloc[h]
    data = df.iloc[h + 1 :].copy().reset_index(drop=True)

    col_map = {}
    for j, v in enumerate(header_row):
        if pd.isna(v): continue
        vu = str(v).upper().strip()
        if "VEHICLE CODE" in vu or "VC" in vu or "SHORT VC" in vu or "SHORT_VC" in vu:
            col_map.setdefault("VC", j)
        if "PLAN" in vu:
            col_map.setdefault("PLAN", j)
        if "COVERAGE" in vu or "FRESH VIN" in vu or vu == "FRESH VIN STOCK":
            col_map["STOCK"] = j
        elif vu == "STOCK" and "STOCK" not in col_map:
            col_map["STOCK"] = j

    vc_col = col_map.get("VC", 1)
    plan_col = col_map.get("PLAN", 4)
    stock_col = col_map.get("STOCK", 9)  # default to index 9 (Col J) which is FRESH VIN STOCK in TCF2

    result_rows = []
    for _, row in data.iterrows():
        vc_val = str(row.iloc[vc_col]).strip() if vc_col < len(row) and pd.notna(row.iloc[vc_col]) else ""
        stock_val = row.iloc[stock_col] if stock_col < len(row) else np.nan
        plan_val = row.iloc[plan_col] if plan_col < len(row) else np.nan

        if is_valid_vc(vc_val):
            s = pd.to_numeric(stock_val, errors="coerce")
            p = pd.to_numeric(plan_val, errors="coerce")
            result_rows.append({
                "SHORT_VC": vc_val,
                "STOCK": s if pd.notna(s) else 0,
                "PLAN": p if pd.notna(p) else 0,
                "LINE": "TCF2",
            })

    return pd.DataFrame(result_rows) if result_rows else pd.DataFrame(
        columns=["SHORT_VC", "STOCK", "PLAN", "LINE"]
    )


@st.cache_data
def parse_cockpit_file(src_type, src_path, sheet_name, line_label, file_mtime: float = 0.0) -> pd.DataFrame:
    """Parse a cockpit / DPT raw file.
    Extracts: Cockpit WH Number, Short VC, Stock/Coverage, Today's O/P.
    """
    if src_type in ["uploaded", "scanned"]:
        engine = "pyxlsb" if str(src_path).lower().endswith(".xlsb") else None
        try:
            with pd.ExcelFile(src_path, engine=engine) as xls:
                sheets = xls.sheet_names
                if sheet_name not in sheets:
                    if "CONSUMPTION" in sheets:
                        sheet_name = "CONSUMPTION"
                    elif "Cockpit" in sheets:
                        sheet_name = "Cockpit"
                    else:
                        sheet_name = sheets[0]
        except Exception:
            pass

    df = read_sheet(src_type, src_path, sheet_name, header=None)

    # Find header row containing "Part No" or "Cockpit WH" or "VC No"
    h = _find_header_row(df, ["PART NO", "COCKPIT WH", "VC NO", "W/H REQUIREMENT", "MODEL", "VEHICLE CODE"])
    if h is None:
        h = 3  # fallback for TCF1 Cockpit format

    header_row = df.iloc[h]
    data = df.iloc[h + 1 :].copy().reset_index(drop=True)
    if data.empty:
        return pd.DataFrame(columns=["COCKPIT_WH", "SHORT_VC", "COVERAGE", "TODAY_OP", "LINE"])

    # Detect column indices from header
    col_map = {}
    for j, v in enumerate(header_row):
        if pd.isna(v):
            continue
        vu = str(v).upper().strip()
        if "PART NO" in vu and "AB12" not in vu and "COCKPIT" not in vu:
            col_map.setdefault("COCKPIT_WH", j)
        if "COCKPIT" in vu and ("WH" in vu or "PART" in vu or "NUMBER" in vu):
            col_map["COCKPIT_WH"] = j
        if vu == "MODEL":
            col_map.setdefault("COCKPIT_WH", j)
            
        if ("VC" in vu or "VEHICLE" in vu) and ("NO" in vu or "CODE" in vu):
            col_map.setdefault("SHORT_VC", j)
        if "SHORT VC" in vu:
            col_map["SHORT_VC"] = j
            
        if "COVERAGE" in vu and "6:30" in vu:
            col_map["COVERAGE"] = j
        if "FLOAT COVERAGE" in vu and "6:30" in vu:
            col_map["COVERAGE"] = j
        if vu == "TOTAL":
            col_map.setdefault("COVERAGE", j)
            
        if "TODAY" in vu and "O/P" in vu:
            col_map.setdefault("TODAY_OP", j)

    # Fallbacks based on known structure
    cockpit_col = col_map.get("COCKPIT_WH", 0)
    vc_col = col_map.get("SHORT_VC", 4)
    coverage_col = col_map.get("COVERAGE", 20)
    today_col = col_map.get("TODAY_OP", 12)

    result_rows = []
    for _, row in data.iterrows():
        cwh = str(row.iloc[cockpit_col]).strip() if pd.notna(row.iloc[cockpit_col]) else ""
        vc = str(row.iloc[vc_col]).strip() if pd.notna(row.iloc[vc_col]) else ""

        if not (len(cwh) >= 6 and cwh[0].isdigit()) and not is_valid_vc(vc):
            continue

        cov = 0
        if coverage_col is not None and coverage_col < len(row):
            cov = pd.to_numeric(row.iloc[coverage_col], errors="coerce")
            cov = cov if pd.notna(cov) else 0

        top = 0
        if today_col < len(row):
            top = pd.to_numeric(row.iloc[today_col], errors="coerce")
            top = top if pd.notna(top) else 0

        result_rows.append({
            "COCKPIT_WH": cwh if len(cwh) >= 6 else np.nan,
            "SHORT_VC": vc,
            "COVERAGE": cov,
            "TODAY_OP": top,
            "LINE": line_label,
        })

    return pd.DataFrame(result_rows) if result_rows else pd.DataFrame(
        columns=["COCKPIT_WH", "SHORT_VC", "COVERAGE", "TODAY_OP", "LINE"]
    )


# ═══════════════════════════════════════════════════════════════
# 5. COMPUTATION FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def compute_model_wise_float(paint_df: pd.DataFrame, dpt_plans: pd.DataFrame | None, line_filter: str) -> pd.DataFrame:
    """Aggregate Paint Float by model and line, append Today VIN."""
    df = paint_df.copy()
    if line_filter != "All Lines":
        df = df[df["LINE"] == line_filter]

    if df.empty:
        return pd.DataFrame()

    agg = {c: "sum" for c in FLOAT_COLS}
    summary = df.groupby(["LINE", "MODEL"], as_index=False).agg(agg)

    if dpt_plans is not None and not dpt_plans.empty:
        dpt = dpt_plans.copy()
        def classify_dpt_vc(vc):
            vc_str = str(vc).strip()[:4]
            for k, (mod, ln) in MODEL_MAP.items():
                if vc_str.startswith(k[:4]):
                    return pd.Series([mod, ln])
            return pd.Series(["OTHER", "UNKNOWN"])
        
        dpt[["MODEL_DPT", "LINE_DPT"]] = dpt["SHORT_VC"].apply(classify_dpt_vc)
        today_vin_df = dpt.groupby(["LINE_DPT", "MODEL_DPT"], as_index=False)["DPT_VIN"].sum().rename(columns={"LINE_DPT": "LINE", "MODEL_DPT": "MODEL", "DPT_VIN": "TODAY_VIN"})
        summary = summary.merge(today_vin_df, on=["LINE", "MODEL"], how="left")
    else:
        summary["TODAY_VIN"] = 0

    summary["TODAY_VIN"] = summary["TODAY_VIN"].fillna(0).astype(int)

    rows_extra = []
    for line in summary["LINE"].unique():
        sub = summary[summary["LINE"] == line]
        total = sub[FLOAT_COLS + ["TODAY_VIN"]].sum()
        total["LINE"] = line
        total["MODEL"] = f"▸ {line} TOTAL"
        rows_extra.append(total)

    grand = summary[FLOAT_COLS + ["TODAY_VIN"]].sum()
    grand["LINE"] = ""
    grand["MODEL"] = "▸▸ GRAND TOTAL"
    rows_extra.append(grand)

    summary = pd.concat([summary, pd.DataFrame(rows_extra)], ignore_index=True)

    display_cols = ["LINE", "MODEL"] + FLOAT_COLS + ["TODAY_VIN"]
    return summary[display_cols].rename(columns={
        "LINE": "Paint Float",
        "MODEL": "MODEL",
        "TOTAL_FLOAT": "TOTAL FLOAT",
        "PBS_FLOAT": "PBS FLOAT",
        "PBS_TO_POLISHING": "PBS TO POLISHING",
        "POLISHING_TO_TOPCOAT": "POLISHING TO TOPCOAT",
        "TOPCOAT_TO_WETSANDING_ROOFBLACK": "TOPCOAT TO WETSANDING ROOFBLACK",
        "TOPCOAT_TO_WETSANDING_FRESH": "TOPCOAT TO WETSANDING FRESH",
        "WETSANDING_TO_SEALANT": "WETSANDING TO SEALANT",
        "TOTAL_UPTO_SEALANT": "TOTAL UPTO SEALANT",
        "PT_ENTRY_TO_SEALENT": "PT ENTRY TO SEALENT",
        "BIW_LIFTING_TO_PT": "BIW LIFTING TO PT",
        "PT_BYPASS": "PT BYPASS",
        "TODAY_VIN": "Today VIN",
    })


def compute_wiring_summary(
    paint_df: pd.DataFrame,
    wiring_tcf1: pd.DataFrame | None,
    wiring_tcf2: pd.DataFrame | None,
    part_master: pd.DataFrame,
    dpt_plans: pd.DataFrame | None,
    line_filter: str,
) -> pd.DataFrame:
    """Compute wiring harness shortage report.

    Float requirements come from Paint Float (grouped by WIRING_PART_NUMBER).
    Stock/clearance come from wiring files or Part Number Master mapping.
    DPT plans provide Today VIN by variant.
    """
    # ── 1. Float requirements from Paint Float ──
    pf = paint_df.copy()
    pf["WIRING_PART_NUMBER"] = pf["WIRING_PART_NUMBER"].astype(str).str.strip()
    pf = pf[pf["WIRING_PART_NUMBER"].notna() & (pf["WIRING_PART_NUMBER"].str.len() >= 10)]

    req = (
        pf.groupby(["WIRING_PART_NUMBER", "LINE"], as_index=False)
        .agg(
            PAINT_TOTAL_FLOAT=("TOTAL_FLOAT", "sum"),
            PBS_FLOAT=("PBS_FLOAT", "sum"),
            CABS_FLOAT_UPTO_SEALANT=("TOTAL_UPTO_SEALANT", "sum"),
        )
    )

    # Also build from Part Number Master for parts not currently in float
    pm_wiring = part_master[["SHORT_VC", "FRONT_WIRING"]].dropna(subset=["FRONT_WIRING"]).drop_duplicates()
    pm_wiring["FRONT_WIRING"] = pm_wiring["FRONT_WIRING"].astype(str).str.strip()
    pm_parts = set(pm_wiring["FRONT_WIRING"].unique())
    existing_parts = set(req["WIRING_PART_NUMBER"].unique())
    missing_parts = pm_parts - existing_parts
    if missing_parts:
        # Determine line for missing parts via PM → Short VC → FOR_MODEL_FLOAT mapping
        extra_rows = []
        for wp in missing_parts:
            vcs = pm_wiring[pm_wiring["FRONT_WIRING"] == wp]["SHORT_VC"].values
            line = "TCF1"  # default
            for vc in vcs:
                prefix = str(vc)[:4]
                for k, (_, ln) in MODEL_MAP.items():
                    if prefix.startswith(k[:4]):
                        line = ln
                        break
            extra_rows.append({
                "WIRING_PART_NUMBER": wp,
                "LINE": line,
                "PAINT_TOTAL_FLOAT": 0,
                "PBS_FLOAT": 0,
                "CABS_FLOAT_UPTO_SEALANT": 0,
            })
        if extra_rows:
            req = pd.concat([req, pd.DataFrame(extra_rows)], ignore_index=True)

    # ── 2. Stock/clearance from wiring files ──
    stock_records = {}  # wiring_part → {STOCK, PLAN}

    # TCF1: has wiring part directly
    if wiring_tcf1 is not None and not wiring_tcf1.empty:
        for _, r in wiring_tcf1.iterrows():
            wp = str(r.get("WIRING_PART", "")).strip()
            if len(wp) >= 10:
                if wp not in stock_records or r["STOCK"] > stock_records[wp]["STOCK"]:
                    stock_records[wp] = {"STOCK": r["STOCK"], "PLAN": r["PLAN"]}

    # TCF2: needs PM mapping Short VC → Front Wiring
    if wiring_tcf2 is not None and not wiring_tcf2.empty:
        tcf2_merged = wiring_tcf2.merge(pm_wiring, on="SHORT_VC", how="left")
        for _, r in tcf2_merged.iterrows():
            wp = str(r.get("FRONT_WIRING", "")).strip()
            if len(wp) >= 10 and wp != "nan":
                if wp not in stock_records or r["STOCK"] > stock_records[wp]["STOCK"]:
                    stock_records[wp] = {"STOCK": r["STOCK"], "PLAN": r["PLAN"]}

    # ── 3. Today VIN from DPT Plans (sum plan per wiring part) ──
    today_vin_map = {}
    if dpt_plans is not None and not dpt_plans.empty:
        dpt = dpt_plans.copy()
        dpt["WIRING"] = dpt["WIRING"].astype(str).str.strip()
        dpt_agg = dpt[dpt["WIRING"].str.len() >= 10].groupby("WIRING", as_index=False)["DPT_VIN"].sum()
        today_vin_map = dict(zip(dpt_agg["WIRING"], dpt_agg["DPT_VIN"]))

    # ── 4. Merge everything ──
    stock_df = pd.DataFrame([
        {"WIRING_PART_NUMBER": k, "CLEARANCE": v["STOCK"], "TODAY_VIN_W": v["PLAN"]}
        for k, v in stock_records.items()
    ])

    if not stock_df.empty:
        summary = req.merge(stock_df, on="WIRING_PART_NUMBER", how="left")
    else:
        summary = req.copy()
        summary["CLEARANCE"] = 0
        summary["TODAY_VIN_W"] = 0

    summary["CLEARANCE"] = summary["CLEARANCE"].fillna(0).astype(int)
    summary["TODAY_VIN_W"] = summary["TODAY_VIN_W"].fillna(0).astype(int)

    # Override Today VIN with DPT plan data if available
    summary["TODAY_VIN"] = summary["WIRING_PART_NUMBER"].map(today_vin_map).fillna(summary["TODAY_VIN_W"]).astype(int)
    summary.drop(columns=["TODAY_VIN_W"], inplace=True, errors="ignore")

    # ── 5. Compute shortage ──
    summary["SHORTAGE_PBS"] = summary["CLEARANCE"] - summary["TODAY_VIN"] - summary["PBS_FLOAT"]
    summary["SHORTAGE_SEALANT"] = summary["CLEARANCE"] - summary["TODAY_VIN"] - summary["CABS_FLOAT_UPTO_SEALANT"]
    summary["SHORTAGE_TOTAL"] = summary["CLEARANCE"] - summary["TODAY_VIN"] - summary["PAINT_TOTAL_FLOAT"]

    # Filter by line
    if line_filter != "All Lines":
        summary = summary[summary["LINE"] == line_filter]

    # Sort: worst shortages first
    summary = summary.sort_values("SHORTAGE_TOTAL", ascending=True).reset_index(drop=True)

    return summary.rename(columns={
        "WIRING_PART_NUMBER": "Wiring Part Number",
        "LINE": "Model/Line",
        "CLEARANCE": "Clearance After 6:30AM",
        "PAINT_TOTAL_FLOAT": "Paint TOTAL FLOAT",
        "PBS_FLOAT": "PBS FLOAT",
        "CABS_FLOAT_UPTO_SEALANT": "Cabs FloatUPTO SEALANT",
        "SHORTAGE_PBS": "Shortage PBS FLOAT",
        "SHORTAGE_SEALANT": "Shortage Upto Sealant",
        "SHORTAGE_TOTAL": "Shortage for TOTAL FLOAT",
    })[["Wiring Part Number", "Model/Line", "Clearance After 6:30AM",
        "Paint TOTAL FLOAT", "PBS FLOAT", "Cabs FloatUPTO SEALANT",
        "Shortage PBS FLOAT", "Shortage Upto Sealant", "Shortage for TOTAL FLOAT"]]


def compute_cockpit_summary(
    paint_df: pd.DataFrame,
    cockpit_dfs: list[pd.DataFrame],
    part_master: pd.DataFrame,
    dpt_plans: pd.DataFrame | None,
    line_filter: str,
) -> pd.DataFrame:
    """Compute cockpit assembly shortage report.

    Float requirements come from Paint Float (grouped by COCKPIT column).
    Stock/coverage comes from cockpit raw files mapped via Part Number Master.
    """
    # ── 1. Float requirements from Paint Float ──
    pf = paint_df.copy()
    pf["COCKPIT"] = pf["COCKPIT"].astype(str).str.strip()
    pf = pf[pf["COCKPIT"].notna() & (pf["COCKPIT"].str.len() >= 10)]

    req = (
        pf.groupby(["COCKPIT", "LINE"], as_index=False)
        .agg(
            PAINT_TOTAL_FLOAT=("TOTAL_FLOAT", "sum"),
            PBS_FLOAT=("PBS_FLOAT", "sum"),
            CABS_FLOAT_UPTO_SEALANT=("TOTAL_UPTO_SEALANT", "sum"),
        )
    )

    # Add PM-only parts not currently in float
    pm_cockpit = part_master[["SHORT_VC", "COCKPIT"]].dropna(subset=["COCKPIT"]).drop_duplicates()
    pm_cockpit["COCKPIT"] = pm_cockpit["COCKPIT"].astype(str).str.strip()
    pm_parts = set(pm_cockpit["COCKPIT"].unique())
    existing = set(req["COCKPIT"].unique())
    for cp in pm_parts - existing:
        vcs = pm_cockpit[pm_cockpit["COCKPIT"] == cp]["SHORT_VC"].values
        line = "TCF1"
        for vc in vcs:
            prefix = str(vc)[:4]
            for k, (_, ln) in MODEL_MAP.items():
                if prefix.startswith(k[:4]):
                    line = ln
                    break
        req = pd.concat([req, pd.DataFrame([{
            "COCKPIT": cp, "LINE": line,
            "PAINT_TOTAL_FLOAT": 0, "PBS_FLOAT": 0, "CABS_FLOAT_UPTO_SEALANT": 0,
        }])], ignore_index=True)

    # ── 2. Stock from cockpit files (mapped via Short VC → PM COCKPIT) ──
    stock_map = {}  # cockpit_module_part → {COVERAGE, TODAY_OP}

    valid_dfs = [d for d in cockpit_dfs if d is not None and not d.empty]
    if valid_dfs:
        combined_cockpit = pd.concat(valid_dfs, ignore_index=True)
    else:
        combined_cockpit = pd.DataFrame()
        
    if not combined_cockpit.empty:
        # Map Short VC → COCKPIT module via PM
        merged = combined_cockpit.merge(pm_cockpit, on="SHORT_VC", how="left", suffixes=("_raw", "_pm"))
        cockpit_col = "COCKPIT_pm" if "COCKPIT_pm" in merged.columns else "COCKPIT"
        merged["COCKPIT_MODULE"] = merged.get(cockpit_col, pd.Series(dtype=str)).astype(str).str.strip()
        valid = merged[merged["COCKPIT_MODULE"].str.len() >= 10]

        if not valid.empty:
            agg = valid.groupby("COCKPIT_MODULE", as_index=False).agg(
                COVERAGE=("COVERAGE", "sum"),
                TODAY_OP=("TODAY_OP", "sum"),
            )
            stock_map = dict(zip(agg["COCKPIT_MODULE"], agg[["COVERAGE", "TODAY_OP"]].to_dict("records")))

    # ── 3. Today VIN from DPT plans ──
    today_vin_map = {}
    if dpt_plans is not None and not dpt_plans.empty:
        dpt = dpt_plans.copy()
        dpt["COCKPIT"] = dpt["COCKPIT"].astype(str).str.strip()
        dpt_agg = dpt[dpt["COCKPIT"].str.len() >= 10].groupby("COCKPIT", as_index=False)["DPT_VIN"].sum()
        today_vin_map = dict(zip(dpt_agg["COCKPIT"], dpt_agg["DPT_VIN"]))

    # ── 4. Merge ──
    req["CLEARANCE"] = req["COCKPIT"].map(lambda x: stock_map.get(x, {}).get("COVERAGE", 0)).fillna(0).astype(int)
    cov_today = req["COCKPIT"].map(lambda x: stock_map.get(x, {}).get("TODAY_OP", 0)).fillna(0).astype(int)
    req["TODAY_VIN"] = req["COCKPIT"].map(today_vin_map).fillna(cov_today).astype(int)

    # ── 5. Shortage ──
    req["SHORTAGE_PBS"] = req["CLEARANCE"] - req["TODAY_VIN"] - req["PBS_FLOAT"]
    req["SHORTAGE_SEALANT"] = req["CLEARANCE"] - req["TODAY_VIN"] - req["CABS_FLOAT_UPTO_SEALANT"]
    req["SHORTAGE_TOTAL"] = req["CLEARANCE"] - req["TODAY_VIN"] - req["PAINT_TOTAL_FLOAT"]

    if line_filter != "All Lines":
        req = req[req["LINE"] == line_filter]

    req = req.sort_values("SHORTAGE_TOTAL", ascending=True).reset_index(drop=True)

    return req.rename(columns={
        "COCKPIT": "Cockpit Part Number",
        "LINE": "Model",
        "CLEARANCE": "Clearance After 6:30AM",
        "PAINT_TOTAL_FLOAT": "Paint TOTAL FLOAT",
        "PBS_FLOAT": "PBS FLOAT",
        "CABS_FLOAT_UPTO_SEALANT": "Cabs Float UPTO SEALANT",
        "SHORTAGE_PBS": "Shortage PBS FLOAT",
        "SHORTAGE_SEALANT": "Shortage Upto Sealant",
        "SHORTAGE_TOTAL": "Shortage TOTAL FLOAT",
    })[["Cockpit Part Number", "Model", "Clearance After 6:30AM",
        "Paint TOTAL FLOAT", "PBS FLOAT", "Cabs Float UPTO SEALANT",
        "Shortage PBS FLOAT", "Shortage Upto Sealant", "Shortage TOTAL FLOAT"]]

def compute_vin_vs_float(
    paint_df: pd.DataFrame,
    dpt_plans: pd.DataFrame | None,
    line_filter: str,
) -> pd.DataFrame:
    """Consolidated VIN Generation vs Paint Float shortage per Short VC.

    Matches the reference workbook intent: each row is one Short VC showing
    DPT Plan, Today VIN, all float stages, and shortage columns.
    """
    if paint_df.empty:
        return pd.DataFrame()

    pf = paint_df.copy()
    if line_filter != "All Lines":
        pf = pf[pf["LINE"] == line_filter]

    if pf.empty:
        return pd.DataFrame()

    # Base: one row per Short VC with all float data
    base = pf[[
        "SHORT_VC", "SALES_DESCRIPTION", "MODEL", "LINE",
        "TOTAL_FLOAT", "PBS_FLOAT", "PBS_TO_POLISHING",
        "POLISHING_TO_TOPCOAT", "TOTAL_UPTO_SEALANT",
        "BIW_LIFTING_TO_PT", "PT_ENTRY_TO_SEALENT",
    ]].copy()

    # Merge DPT VIN per Short VC
    if dpt_plans is not None and not dpt_plans.empty:
        dpt = dpt_plans[["SHORT_VC", "DPT_PLAN", "DPT_VIN"]].copy()
        dpt["SHORT_VC"] = dpt["SHORT_VC"].astype(str).str.strip()
        dpt = dpt.groupby("SHORT_VC", as_index=False).agg(
            DPT_PLAN=("DPT_PLAN", "sum"),
            DPT_VIN=("DPT_VIN", "sum"),
        )
        base = base.merge(dpt, on="SHORT_VC", how="left")
    else:
        base["DPT_PLAN"] = 0
        base["DPT_VIN"] = 0

    base["DPT_PLAN"] = base["DPT_PLAN"].fillna(0).astype(int)
    base["DPT_VIN"]  = base["DPT_VIN"].fillna(0).astype(int)

    # Shortage = Today VIN (DPT_VIN) – Float
    base["Shortage vs PBS Float"]        = base["DPT_VIN"] - base["PBS_FLOAT"]
    base["Shortage vs Upto Sealant"]     = base["DPT_VIN"] - base["TOTAL_UPTO_SEALANT"]
    base["Shortage vs TOTAL Float"]      = base["DPT_VIN"] - base["TOTAL_FLOAT"]

    # Sort: worst shortage first, then by model
    base = base.sort_values(
        ["Shortage vs TOTAL Float", "MODEL", "SHORT_VC"],
        ascending=[True, True, True]
    ).reset_index(drop=True)

    return base.rename(columns={
        "SHORT_VC":            "Short VC",
        "SALES_DESCRIPTION":   "Sales Description",
        "MODEL":               "Model",
        "LINE":                "Line",
        "DPT_PLAN":            "DPT Plan",
        "DPT_VIN":             "Today VIN (DPT)",
        "TOTAL_FLOAT":         "TOTAL Float",
        "PBS_FLOAT":           "PBS Float",
        "PBS_TO_POLISHING":    "PBS→Polishing",
        "POLISHING_TO_TOPCOAT":"Polishing→Topcoat",
        "TOTAL_UPTO_SEALANT":  "Upto Sealant",
        "BIW_LIFTING_TO_PT":   "BIW→PT",
        "PT_ENTRY_TO_SEALENT": "PT→Sealant",
    })[["Short VC", "Sales Description", "Model", "Line",
        "DPT Plan", "Today VIN (DPT)",
        "TOTAL Float", "PBS Float", "Upto Sealant", "BIW→PT", "PT→Sealant",
        "Shortage vs PBS Float", "Shortage vs Upto Sealant", "Shortage vs TOTAL Float"]]

def build_engine_table(paint_df: pd.DataFrame, line_filter: str) -> pd.DataFrame:
    """Build engine requirements table combining Paint Float logic with predefined Master list."""
    # 1. Initialize master df
    req = pd.DataFrame(ENGINE_MASTER)
    
    # 2. Group current paint float — group by ENGINE + LINE only to avoid duplicates
    pf = paint_df.copy()
    pf["ENGINE"] = pf["ENGINE"].astype(str).str.strip()
    pf = pf[pf["ENGINE"].notna() & (pf["ENGINE"].str.len() >= 5)]
    
    pf_agg = (
        pf.groupby(["ENGINE", "LINE"], as_index=False)
        .agg(
            TOTAL_FLOAT=("TOTAL_FLOAT", "sum"),
            PBS_FLOAT=("PBS_FLOAT", "sum"),
            UPTO_SEALANT=("TOTAL_UPTO_SEALANT", "sum"),
        )
    )
    
    # 3. Outer merge to capture all master engines + any new ones
    req = pd.merge(req, pf_agg, left_on=["Engine Part No", "LINE"], right_on=["ENGINE", "LINE"], how="outer", suffixes=("", "_pf"))
    
    # 4. Handle any missing data for newly discovered engines
    missing_mask = req["Engine Part No"].isna()
    req.loc[missing_mask, "Engine Part No"] = req.loc[missing_mask, "ENGINE"]
    req["Model"] = req["Model"].fillna("NEW_ENGINE")
    req["TA Code"] = req["TA Code"].fillna("")
    
    req["TOTAL_FLOAT"] = req["TOTAL_FLOAT"].fillna(0).astype(int)
    req["PBS_FLOAT"] = req["PBS_FLOAT"].fillna(0).astype(int)
    req["UPTO_SEALANT"] = req["UPTO_SEALANT"].fillna(0).astype(int)
    
    req = req.drop(columns=["ENGINE"], errors="ignore")
    
    if line_filter != "All Lines":
        req = req[req["LINE"] == line_filter]

    return req


def _build_engine_manual_display(df: pd.DataFrame) -> pd.DataFrame:
    """Build the manual entry table with subtotal rows for the 4-column editor.
    
    Organizes engines by line (TCF1/TCF2) with subtotal rows:
    - 1.2 Lit engines → '1.2 Lit Total' subtotal
    - Nova
    - 'TCF1' total
    - 2 Lit engines → '2 Lit Total' subtotal
    - Harrier EV
    - 'TCF2' total
    """
    rows = []
    num_cols = ["Clearance After 6:30AM"]

    tcf1_order = ["Punch MT SA", "Punch AMT SA", "Punch TC MCE", "Punch MCE MT", "Punch MCE AMT", "Punch MCE CNG MT", "Punch MCE CNG AMT"]
    tcf2_order = ["Harrier / Safari Diesel AT", "Harrier / Safari Diesel MT", "Harrier / Safari Petrol TGDI MT", "Harrier / Safari Petrol TGDI AT"]

    tcf1_df = df[df["LINE"] == "TCF1"].copy()
    if not tcf1_df.empty:
        is_punch = tcf1_df["Model"].str.contains("PUNCH", case=False, na=False)
        punch_df = tcf1_df[is_punch].copy()
        other_df = tcf1_df[~is_punch].copy()

        punch_df['SortKey'] = pd.Categorical(punch_df['Model'], categories=tcf1_order, ordered=True)
        punch_df = punch_df.sort_values('SortKey').drop(columns=['SortKey'])

        rows.extend(punch_df[["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM"]].to_dict('records'))
        if not punch_df.empty:
            sub = {"Engine Part No": "", "Model": "1.2 Lit Total", "TA Code": "", "Clearance After 6:30AM": punch_df["Clearance After 6:30AM"].sum()}
            rows.append(sub)

        rows.extend(other_df[["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM"]].to_dict('records'))

        tcf1_sub = {"Engine Part No": "", "Model": "TCF1", "TA Code": "", "Clearance After 6:30AM": tcf1_df["Clearance After 6:30AM"].sum()}
        rows.append(tcf1_sub)

    tcf2_df = df[df["LINE"] == "TCF2"].copy()
    if not tcf2_df.empty:
        is_2l = tcf2_df["Model"].str.contains("HARRIER", case=False, na=False) | tcf2_df["Model"].str.contains("SAFARI", case=False, na=False)
        is_ev = tcf2_df["Model"].str.contains("EV", case=False, na=False)
        is_2l = is_2l & ~is_ev

        l2_df = tcf2_df[is_2l].copy()
        ev_df = tcf2_df[~is_2l].copy()

        l2_df['SortKey'] = pd.Categorical(l2_df['Model'], categories=tcf2_order, ordered=True)
        l2_df = l2_df.sort_values('SortKey').drop(columns=['SortKey'])

        rows.extend(l2_df[["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM"]].to_dict('records'))
        if not l2_df.empty:
            sub = {"Engine Part No": "", "Model": "2 Lit Total", "TA Code": "", "Clearance After 6:30AM": l2_df["Clearance After 6:30AM"].sum()}
            rows.append(sub)

        rows.extend(ev_df[["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM"]].to_dict('records'))

        tcf2_sub = {"Engine Part No": "", "Model": "TCF2", "TA Code": "", "Clearance After 6:30AM": tcf2_df["Clearance After 6:30AM"].sum()}
        rows.append(tcf2_sub)

    return pd.DataFrame(rows)


def add_engine_subtotals(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    num_cols = ["Clearance After 6:30AM", "Today VIN", "Bal", "PBS FLOAT", "Float UPTO SEALANT", "TOTAL FLOAT", "With respect to PBS FLOAT", "With respect to Sealant FLOAT", "With respect to Total FLOAT"]

    tcf1_order = ["Punch MT SA", "Punch AMT SA", "Punch TC MCE", "Punch MCE MT", "Punch MCE AMT", "Punch MCE CNG MT", "Punch MCE CNG AMT"]
    tcf2_order = ["Harrier / Safari Diesel AT", "Harrier / Safari Diesel MT", "Harrier / Safari Petrol TGDI MT", "Harrier / Safari Petrol TGDI AT"]

    tcf1_df = df[df["LINE"] == "TCF1"].copy()
    if not tcf1_df.empty:
        is_punch = tcf1_df["Model"].str.contains("PUNCH", case=False, na=False)
        punch_df = tcf1_df[is_punch].copy()
        other_df = tcf1_df[~is_punch].copy()
        
        punch_df['SortKey'] = pd.Categorical(punch_df['Model'], categories=tcf1_order, ordered=True)
        punch_df = punch_df.sort_values('SortKey').drop(columns=['SortKey'])

        rows.extend(punch_df.to_dict('records'))
        if not punch_df.empty:
            sub = {c: punch_df[c].sum() if c in num_cols else "" for c in df.columns}
            sub["Model"] = "1.2 Lit Total"
            sub["LINE"] = "TCF1"
            rows.append(sub)

        rows.extend(other_df.to_dict('records'))

        tcf1_sub = {c: tcf1_df[c].sum() if c in num_cols else "" for c in df.columns}
        tcf1_sub["Model"] = "TCF1"
        tcf1_sub["LINE"] = "TCF1"
        rows.append(tcf1_sub)

    tcf2_df = df[df["LINE"] == "TCF2"].copy()
    if not tcf2_df.empty:
        is_2l = tcf2_df["Model"].str.contains("HARRIER", case=False, na=False) | tcf2_df["Model"].str.contains("SAFARI", case=False, na=False)
        is_ev = tcf2_df["Model"].str.contains("EV", case=False, na=False)
        is_2l = is_2l & ~is_ev

        l2_df = tcf2_df[is_2l].copy()
        ev_df = tcf2_df[~is_2l].copy()
        
        l2_df['SortKey'] = pd.Categorical(l2_df['Model'], categories=tcf2_order, ordered=True)
        l2_df = l2_df.sort_values('SortKey').drop(columns=['SortKey'])

        rows.extend(l2_df.to_dict('records'))
        if not l2_df.empty:
            sub = {c: l2_df[c].sum() if c in num_cols else "" for c in df.columns}
            sub["Model"] = "2 Lit Total"
            sub["LINE"] = "TCF2"
            rows.append(sub)
            
        rows.extend(ev_df.to_dict('records'))
        
        tcf2_sub = {c: tcf2_df[c].sum() if c in num_cols else "" for c in df.columns}
        tcf2_sub["Model"] = "TCF2"
        tcf2_sub["LINE"] = "TCF2"
        rows.append(tcf2_sub)
        
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════
# 6. STYLING FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def highlight_shortage(val):
    """Traffic-light styling for shortage columns."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return ""
    if v < 0:
        return "background-color: #ffe4e6; color: #9f1239; font-weight: 700; border: 1px solid #fecaca;"
    if v == 0:
        return "background-color: #fef3c7; color: #b45309; font-weight: 600; border: 1px solid #fde68a;"
    return ""


def style_shortage_df(df: pd.DataFrame, table_type="wiring"):
    """Apply conditional formatting to shortage columns in a dataframe."""
    shortage_cols = [c for c in df.columns if "Shortage" in str(c) or "With respect to" in str(c)]
    
    styler = df.style
    
    styles = {}
    for i, col in enumerate(df.columns):
        if table_type in ["wiring", "cockpit", "engine"]:
            bg_color = '#475569' if i < 3 else '#1e3a8a'
            text_color = '#ffffff'
        else:
            bg_color = '#334155'
            text_color = '#ffffff'
        styles[col] = [{'selector': 'th', 'props': [('background-color', bg_color), ('color', text_color), ('font-weight', '700'), ('border', '1px solid #e2e8f0'), ('padding', '10px 14px')]}]
    
    styler = styler.set_table_styles(styles, overwrite=False)
    
    if shortage_cols:
        styler = styler.map(highlight_shortage, subset=shortage_cols)
        
    return styler


def render_html_table(df: pd.DataFrame) -> str:
    """Build premium HTML table for Model Wise Float summary with wrapped headers and fit-to-page styling."""
    html = '<table class="model-wise-table">'
    
    # Headers
    html += '<thead><tr>'
    for col in df.columns:
        wrapped_col = str(col)
        # Wrap long column names to narrower lines
        if "TO WETSANDING" in wrapped_col:
            wrapped_col = wrapped_col.replace("TO WETSANDING ", "TO<br>WETSANDING<br>")
        elif "WETSANDING TO" in wrapped_col:
            wrapped_col = wrapped_col.replace("WETSANDING TO ", "WETSANDING<br>TO<br>")
        elif "POLISHING TO" in wrapped_col:
            wrapped_col = wrapped_col.replace("POLISHING TO ", "POLISHING<br>TO<br>")
        elif "PBS TO" in wrapped_col:
            wrapped_col = wrapped_col.replace("PBS TO ", "PBS<br>TO<br>")
        elif "TOTAL UPTO" in wrapped_col:
            wrapped_col = wrapped_col.replace("TOTAL UPTO ", "TOTAL<br>UPTO<br>")
        elif "ENTRY TO" in wrapped_col:
            wrapped_col = wrapped_col.replace("ENTRY TO ", "ENTRY<br>TO<br>")
        elif "LIFTING TO" in wrapped_col:
            wrapped_col = wrapped_col.replace("LIFTING TO ", "LIFTING<br>TO<br>")
        elif "TOTAL FLOAT" in wrapped_col:
            wrapped_col = wrapped_col.replace("TOTAL FLOAT", "TOTAL<br>FLOAT")
        elif "PBS FLOAT" in wrapped_col:
            wrapped_col = wrapped_col.replace("PBS FLOAT", "PBS<br>FLOAT")
        elif "Today VIN" in wrapped_col:
            wrapped_col = wrapped_col.replace("Today VIN", "Today<br>VIN")
        elif "Paint Float" in wrapped_col:
            wrapped_col = wrapped_col.replace("Paint Float", "Paint<br>Float")
        
        html += f'<th>{wrapped_col}</th>'
    html += '</tr></thead>'
    
    # Body
    html += '<tbody>'
    for _, row in df.iterrows():
        model_val = str(row.get("MODEL", "")).upper()
        if "GRAND TOTAL" in model_val:
            row_class = 'class="grand-total"'
        elif "TOTAL" in model_val:
            row_class = 'class="line-total"'
        else:
            row_class = ''
            
        html += f'<tr {row_class}>'
        for col in df.columns:
            val = row[col]
            if isinstance(val, (int, float)):
                if pd.isna(val):
                    val_str = ""
                elif val == int(val):
                    val_str = f"{int(val):,}"
                else:
                    val_str = f"{val:,.1f}"
            else:
                val_str = str(val) if not pd.isna(val) else ""
            html += f'<td>{val_str}</td>'
        html += '</tr>'
    html += '</tbody>'
    
    html += '</table>'
    return html


def render_shortage_html_table(df: pd.DataFrame, table_type="wiring") -> str:
    """Build premium HTML table for Wiring and Cockpit shortage reports with peach/blue headers, wrapped text, and double vertical divider lines."""
    plot_df = df.copy()
    
    if "Today VIN" in plot_df.columns:
        plot_df = plot_df.drop(columns=["Today VIN"], errors="ignore")
        
    if table_type == "cockpit":
        plot_df = plot_df.rename(columns={
            "Model/Line": "Model",
            "Cabs FloatUPTO SEALANT": "Cabs Float UPTO SEALANT",
            "Shortage for TOTAL FLOAT": "Shortage TOTAL FLOAT"
        })

    # Wrap headers
    wrapped_headers = []
    for c in plot_df.columns:
        col_name = str(c)
        if "Part Number" in col_name:
            wrapped_headers.append(col_name.replace(" Part Number", "<br>Part Number"))
        elif "Clearance After" in col_name:
            wrapped_headers.append("Clearance<br>After<br>6:30AM")
        elif "Paint TOTAL" in col_name:
            wrapped_headers.append("Paint TOTAL<br>FLOAT")
        elif "PBS FLOAT" in col_name:
            wrapped_headers.append("PBS<br>FLOAT")
        elif "Cabs Float" in col_name:
            wrapped_headers.append("Cabs Float<br>UPTO<br>SEALANT")
        elif "Shortage PBS" in col_name:
            wrapped_headers.append("Shortage PBS<br>FLOAT")
        elif "Upto Sealant" in col_name:
            wrapped_headers.append("Shortage<br>Upto<br>Sealant")
        elif "TOTAL FLOAT" in col_name:
            if "Shortage for" in col_name:
                wrapped_headers.append("Shortage<br>for TOTAL<br>FLOAT")
            else:
                wrapped_headers.append("Shortage<br>TOTAL<br>FLOAT")
        else:
            wrapped_headers.append(col_name)

    shortage_cols = [c for c in plot_df.columns if "Shortage" in str(c)]

    html = '<div style="width: 100%; overflow-x: auto; border-radius: 12px; border: 1px solid #cbd5e1;">'
    html += '<table class="model-wise-table">'
    
    # Headers
    html += '<thead><tr>'
    for idx, col in enumerate(wrapped_headers):
        # Peach for first 3 columns, blue for remaining
        bg_color = "#F8CBAD" if idx < 3 else "#BDD7EE"
        text_color = "#000000"
        
        # Double right border for columns 3 and 6
        border_right = "3px double #000000" if idx in [2, 5] else "1px solid #cbd5e1"
        
        html += f'<th style="background-color: {bg_color} !important; color: {text_color} !important; border-right: {border_right} !important;">{col}</th>'
    html += '</tr></thead>'
    
    # Body
    html += '<tbody>'
    for _, row in plot_df.iterrows():
        html += '<tr>'
        for idx, col in enumerate(plot_df.columns):
            val = row[col]
            bg = "#ffffff"
            tc = "#000000"
            font_weight = "normal"
            
            if isinstance(val, (int, float)):
                if pd.isna(val):
                    val_str = ""
                elif val == int(val):
                    val_str = f"{int(val):,}"
                else:
                    val_str = f"{val:,.1f}"
                
                if col in shortage_cols and val < 0:
                    bg = "#FCE4D6"
                    tc = "#C00000"
                    font_weight = "bold"
            else:
                val_str = str(val) if not pd.isna(val) else ""
            
            border_right = "3px double #000000" if idx in [2, 5] else "1px solid #cbd5e1"
            
            html += f'<td style="background-color: {bg} !important; color: {tc} !important; font-weight: {font_weight} !important; border-right: {border_right} !important;">{val_str}</td>'
        html += '</tr>'
    html += '</tbody>'
    html += '</table></div>'
    return html


def render_engine_html_table(df: pd.DataFrame) -> str:
    """Build premium HTML table for Engine & Battery summary with custom row grouping colors and highlights."""
    plot_df = df.copy()
    
    if "LINE" in plot_df.columns:
        plot_df = plot_df.drop(columns=["LINE"], errors="ignore")

    wrapped_headers = []
    for c in plot_df.columns:
        col_name = str(c)
        if "Part No" in col_name:
            wrapped_headers.append("Engine<br>Part No")
        elif "Clearance After" in col_name:
            wrapped_headers.append("Clearance<br>After<br>6:30AM")
        elif "Today VIN" in col_name:
            wrapped_headers.append("Today<br>VIN")
        elif "PBS FLOAT" in col_name:
            wrapped_headers.append("PBS<br>FLOAT")
        elif "UPTO SEALANT" in col_name:
            wrapped_headers.append("Float UPTO<br>SEALANT")
        elif "TOTAL FLOAT" in col_name:
            wrapped_headers.append("TOTAL<br>FLOAT")
        elif "With respect to" in col_name:
            wrapped_headers.append(col_name.replace("With respect to ", "w.r.t.<br>"))
        else:
            wrapped_headers.append(col_name)

    shortage_cols = [c for c in plot_df.columns if "respect to" in str(c)]

    html = '<div style="width: 100%; overflow-x: auto; border-radius: 12px; border: 1px solid #cbd5e1;">'
    html += '<table class="model-wise-table">'
    
    # Headers
    html += '<thead><tr>'
    for col in wrapped_headers:
        html += f'<th>{col}</th>'
    html += '</tr></thead>'
    
    # Body
    html += '<tbody>'
    for _, row in plot_df.iterrows():
        model_val = str(row.get("Model", ""))
        
        if "Total" in model_val:
            row_style = 'style="background-color: #e2f0d9 !important; font-weight: 700 !important; color: #1e3a8a !important;"'
        elif model_val in ["TCF1", "TCF2"]:
            row_style = 'style="background-color: #fff2cc !important; font-weight: 800 !important; color: #b45309 !important;"'
        elif "GRAND TOTAL" in model_val.upper() or "GRAND" in model_val.upper():
            row_style = 'style="background-color: #dcfce7 !important; font-weight: 800 !important; color: #0f5132 !important;"'
        else:
            row_style = ''
            
        html += f'<tr {row_style}>'
        for col in plot_df.columns:
            val = row[col]
            bg = ""
            tc = ""
            font_weight = ""
            
            if isinstance(val, (int, float)):
                if pd.isna(val):
                    val_str = ""
                elif val == int(val):
                    val_str = f"{int(val):,}"
                else:
                    val_str = f"{val:,.1f}"
                
                if col in shortage_cols and val < 0 and "Total" not in model_val and model_val not in ["TCF1", "TCF2"]:
                    bg = "background-color: #FCE4D6 !important;"
                    tc = "color: #C00000 !important;"
                    font_weight = "font-weight: bold !important;"
            else:
                val_str = str(val) if not pd.isna(val) else ""
                
            cell_style = f'style="{bg} {tc} {font_weight}"' if (bg or tc or font_weight) else ''
            html += f'<td {cell_style}>{val_str}</td>'
        html += '</tr>'
    html += '</tbody>'
    html += '</table></div>'
    return html


def render_vin_float_html_table(df: pd.DataFrame) -> str:
    """Build premium HTML table for VIN vs Paint Float shortages with styled headers and status highlights."""
    plot_df = df.copy()

    wrapped_headers = []
    for c in plot_df.columns:
        col_name = str(c)
        if "Short VC" in col_name:
            wrapped_headers.append("Short<br>VC")
        elif "Sales Description" in col_name:
            wrapped_headers.append("Sales<br>Description")
        elif "DPT Plan" in col_name:
            wrapped_headers.append("DPT<br>Plan")
        elif "Today VIN" in col_name:
            wrapped_headers.append("Today VIN<br>(DPT)")
        elif "TOTAL Float" in col_name:
            wrapped_headers.append("TOTAL<br>Float")
        elif "PBS Float" in col_name:
            wrapped_headers.append("PBS<br>Float")
        elif "Upto Sealant" in col_name:
            wrapped_headers.append("Upto<br>Sealant")
        elif "BIW" in col_name:
            wrapped_headers.append("BIW→PT")
        elif "PT" in col_name and "?" in col_name:
            wrapped_headers.append("PT→Sealant")
        elif "Shortage vs" in col_name:
            wrapped_headers.append(col_name.replace("Shortage vs ", "Shortage vs<br>"))
        else:
            wrapped_headers.append(col_name)

    shortage_cols = [c for c in plot_df.columns if "Shortage" in str(c)]

    html = '<div style="width: 100%; overflow-x: auto; border-radius: 12px; border: 1px solid #cbd5e1;">'
    html += '<table class="model-wise-table">'
    
    # Headers
    html += '<thead><tr>'
    for col in wrapped_headers:
        html += f'<th>{col}</th>'
    html += '</tr></thead>'
    
    # Body
    html += '<tbody>'
    for _, row in plot_df.iterrows():
        html += '<tr>'
        for col in plot_df.columns:
            val = row[col]
            bg = ""
            tc = ""
            font_weight = ""
            
            if isinstance(val, (int, float)):
                if pd.isna(val):
                    val_str = ""
                elif val == int(val):
                    val_str = f"{int(val):,}"
                else:
                    val_str = f"{val:,.1f}"
                
                if col in shortage_cols:
                    if val < 0:
                        bg = "background-color: #fee2e2 !important;"
                        tc = "color: #991b1b !important;"
                        font_weight = "font-weight: 700 !important;"
                    elif val == 0:
                        bg = "background-color: #fef3c7 !important;"
                        tc = "color: #b45309 !important;"
                        font_weight = "font-weight: 600 !important;"
                    else:
                        bg = "background-color: #dcfce7 !important;"
                        tc = "color: #166534 !important;"
                        
                if col == "Today VIN (DPT)":
                    bg = "background-color: #dbeafe !important;"
                    tc = "color: #1d4ed8 !important;"
                    font_weight = "font-weight: 600 !important;"
            else:
                val_str = str(val) if not pd.isna(val) else ""
                
            cell_style = f'style="{bg} {tc} {font_weight}"' if (bg or tc or font_weight) else ''
            html += f'<td {cell_style}>{val_str}</td>'
        html += '</tr>'
    html += '</tbody>'
    html += '</table></div>'
    return html


def render_vin_float_summary_html_table(df: pd.DataFrame) -> str:
    """Build premium HTML table for Model-Wise summary in VIN vs Float tab."""
    plot_df = df.copy()

    wrapped_headers = []
    for c in plot_df.columns:
        col_name = str(c)
        if "DPT Plan" in col_name:
            wrapped_headers.append("DPT<br>Plan")
        elif "Today VIN" in col_name:
            wrapped_headers.append("Today<br>VIN")
        elif "TOTAL Float" in col_name:
            wrapped_headers.append("TOTAL<br>Float")
        elif "PBS Float" in col_name:
            wrapped_headers.append("PBS<br>Float")
        elif "Upto Sealant" in col_name:
            wrapped_headers.append("Upto<br>Sealant")
        elif "Shortage vs" in col_name:
            wrapped_headers.append(col_name.replace("Shortage vs ", "Shortage vs<br>"))
        else:
            wrapped_headers.append(col_name)

    html = '<div style="width: 100%; overflow-x: auto; border-radius: 12px; border: 1px solid #cbd5e1;">'
    html += '<table class="model-wise-table">'
    
    # Headers
    html += '<thead><tr>'
    for col in wrapped_headers:
        html += f'<th>{col}</th>'
    html += '</tr></thead>'
    
    # Body
    html += '<tbody>'
    for _, row in plot_df.iterrows():
        html += '<tr>'
        for col in plot_df.columns:
            val = row[col]
            if isinstance(val, (int, float)):
                if pd.isna(val):
                    val_str = ""
                elif val == int(val):
                    val_str = f"{int(val):,}"
                else:
                    val_str = f"{val:,.1f}"
            else:
                val_str = str(val) if not pd.isna(val) else ""
            html += f'<td>{val_str}</td>'
        html += '</tr>'
    html += '</tbody>'
    html += '</table></div>'
    return html


# ═══════════════════════════════════════════════════════════════
# 7. MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="TCF PPC Dashboard",
        page_icon="🏭",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    # Debug log session state
    try:
        with open("data/debug_state.log", "a", encoding="utf-8") as f:
            f.write(f"--- RERUN {datetime.now().isoformat()} ---\n")
            f.write(f"session_state keys: {list(st.session_state.keys())}\n")
            for k in list(st.session_state.keys()):
                val = st.session_state[k]
                if isinstance(val, (str, int, float, bool, list, dict)):
                    f.write(f"  {k}: {val}\n")
                else:
                    f.write(f"  {k}: {type(val)} (non-serializable)\n")
    except Exception:
        pass

    # ── Daily 6:30 AM Reset ──
    reset_performed = check_and_perform_daily_reset()
    if reset_performed:
        if "engine_editor_df" in st.session_state:
            del st.session_state["engine_editor_df"]
        st.toast("🔄 Daily 6:30 AM reset completed — all uploaded data cleared.", icon="🔄")
        st.rerun()

    now = get_ist_now()
    file_meta = _load_file_meta()

    # (removed early upload handler; moved processing to tab6 after file_uploader rendering)

    if "show_upload_toast" in st.session_state:
        st.toast(st.session_state["show_upload_toast"], icon="✅")
        del st.session_state["show_upload_toast"]

    # ═══════════════════════════════════════════
    # SIDEBAR
    # ═══════════════════════════════════════════
    with st.sidebar:
        st.markdown("## 🏭 TCF PPC Dashboard")
        st.caption("Production Planning & Control")

        st.divider()

        # ── Filters ──
        st.markdown("### 📊 Filters")
        line_filter = st.selectbox(
            "Main Line",
            ["All Lines", "TCF1", "TCF2"],
            help="Filter the entire dashboard by production line.",
        )
        lookup = st.text_input(
            "🔍 Global VIN / Part Lookup",
            placeholder="e.g. 54680124A or 546854600108",
        )

        pass



    # ═══════════════════════════════════════════
    # LOAD ALL DATA
    # ═══════════════════════════════════════════
    
    # ── BOM Master Data (sole source for part mappings) ──
    bom_df = load_bom_master_cached()
    bom_loaded = not bom_df.empty
    paint_loaded = False
    
    paint_df = pd.DataFrame()
    dpt_all = None
    wiring_tcf1 = None
    wiring_tcf2 = None
    cockpit_dfs = []
    pm = None
    missing_bom_vcs = []

    if bom_loaded:
        # Part master = BOM data directly
        pm = bom_df.drop_duplicates(subset="SHORT_VC").reset_index(drop=True)

        # ── Paint Float ──
        pf_src, pf_path = get_source("Paint Float")
        if pf_src is not None:
            paint_df = load_with_persistent_cache(
                "paint_float",
                str(pf_path),
                parse_paint_float,
                pf_src, str(pf_path), pm, file_mtime=os.path.getmtime(pf_path)
            )
            if not paint_df.empty:
                paint_loaded = True

        if paint_loaded:
            # ── DPT Plans ──
            dpt_frames = []
            for ft, sheet, label in [("TCF1 DPT Plan", "TCF1 DPT Plan", "TCF1"), ("TCF2 DPT Plan", "TCF2 DPT Plan", "TCF2")]:
                s, p = get_source(ft)
                if s:
                    try:
                        dpt_frames.append(
                            load_with_persistent_cache(
                                f"dpt_plan_{label}",
                                str(p),
                                parse_dpt_plan,
                                s, str(p), sheet, label, file_mtime=os.path.getmtime(p)
                            )
                        )
                    except Exception:
                        pass
            dpt_all = pd.concat(dpt_frames, ignore_index=True) if dpt_frames else None
            if dpt_all is not None and not dpt_all.empty:
                # Map WIRING, COCKPIT, and ENGINE columns from Part Master (pm)
                dpt_all.drop(columns=["WIRING", "COCKPIT", "ENGINE"], inplace=True, errors="ignore")
                pm_map = pm[["SHORT_VC", "FRONT_WIRING", "COCKPIT", "ENGINE"]].rename(columns={"FRONT_WIRING": "WIRING"})
                dpt_all = dpt_all.merge(pm_map, on="SHORT_VC", how="left")

            # ── Wiring Files ──
            w1_src, w1_path = get_source("TCF1 Wiring File")
            wiring_tcf1 = load_with_persistent_cache(
                "wiring_tcf1",
                str(w1_path),
                parse_wiring_tcf1,
                w1_src, str(w1_path), file_mtime=os.path.getmtime(w1_path)
            ) if w1_src else None

            w2_src, w2_path = get_source("TCF2 Wiring File")
            wiring_tcf2 = load_with_persistent_cache(
                "wiring_tcf2",
                str(w2_path),
                parse_wiring_tcf2,
                w2_src, str(w2_path), file_mtime=os.path.getmtime(w2_path)
            ) if w2_src else None

            # ── Cockpit Files ──
            for ft, sheet, label in [
                ("TCF1 Cockpit", "TCF1 Cockpit", "TCF1"),
                ("TCF2 Cockpit", "TCF2 Cockpit", "TCF2"),
                ("Nova Cockpit", "Nova Cockpit", "TCF1"),
            ]:
                s, p = get_source(ft)
                if s:
                    try:
                        cockpit_dfs.append(
                            load_with_persistent_cache(
                                f"cockpit_{_safe(ft)}",
                                str(p),
                                parse_cockpit_file,
                                s, str(p), sheet, label, file_mtime=os.path.getmtime(p)
                            )
                        )
                    except Exception:
                        pass

            # Compute BOM mapping integrity issues
            bom_vcs = set(bom_df["SHORT_VC"].dropna().astype(str).str.strip().unique())
            active_vcs = set()
            if paint_loaded and not paint_df.empty:
                active_vcs.update(paint_df["SHORT_VC"].dropna().astype(str).str.strip().unique())
            if dpt_all is not None and not dpt_all.empty:
                active_vcs.update(dpt_all["SHORT_VC"].dropna().astype(str).str.strip().unique())
            missing_bom_vcs = sorted(list(active_vcs - bom_vcs))

    # ═══════════════════════════════════════════
    # HEADER & LIVE STATUS BADGES
    # ═══════════════════════════════════════════
    active_files_count = 0
    for ft in RAW_FILE_TYPES:
        src_type, _ = get_source(ft)
        if src_type is not None:
            active_files_count += 1

    status_badges = []
    # Badge 1: Line Selection
    status_badges.append(
        f'<span class="status-badge status-badge-blue">📍 {line_filter}</span>'
    )
    # Badge 2: Data channel status
    if active_files_count == len(RAW_FILE_TYPES):
        status_badges.append(
            '<span class="status-badge status-badge-green">● All Channels Active (9/9)</span>'
        )
    elif active_files_count > 0:
        status_badges.append(
            f'<span class="status-badge status-badge-amber">● Active Channels ({active_files_count}/9)</span>'
        )
    else:
        status_badges.append(
            '<span class="status-badge status-badge-amber">● No Data Channels</span>'
        )

    # Badge 3: Reset Countdown
    next_reset = now.replace(hour=RESET_HOUR, minute=RESET_MINUTE, second=0, microsecond=0)
    if now >= next_reset:
        next_reset = next_reset + timedelta(days=1)
    mins_to_reset = int((next_reset - now).total_seconds() / 60)
    status_badges.append(
        f'<span class="status-badge">⏱️ Daily Reset in {mins_to_reset} min</span>'
    )

    status_badges_html = f'<div class="status-container">{" ".join(status_badges)}</div>'

    st.markdown(
        '<div class="dash-header">'
        "<h1>🏭 TCF Production Planning Dashboard</h1>"
        f"<p>Live shortage monitoring &nbsp;·&nbsp; Data as of {get_ist_now().strftime('%d %b %Y, %H:%M')}</p>"
        f"{status_badges_html}"
        "</div>",
        unsafe_allow_html=True,
    )

    # ── Global Lookup ──
    if lookup:
        if not bom_loaded:
            st.error("⚠️ **BOM details.xlsx** is not loaded. Cannot perform global search.")
        elif not paint_loaded:
            st.error("⚠️ **Paint Float** is not loaded. Cannot perform global search.")
        else:
            lookup_u = lookup.strip().upper()
            mask = paint_df.apply(lambda row: any(lookup_u in str(v).upper() for v in row), axis=1)
            hits = mask.sum()
            if hits > 0:
                st.info(f"🔍 **{hits}** variant(s) matched for `{lookup}`")
                match_df = paint_df[mask][["SHORT_VC", "SALES_DESCRIPTION", "MODEL", "LINE", "WIRING_PART_NUMBER", "COCKPIT", "ENGINE"] + FLOAT_COLS]
                match_df = match_df.rename(columns={"SHORT_VC": "Short VC", "SALES_DESCRIPTION": "Description", "WIRING_PART_NUMBER": "Wiring Part", "MODEL": "Model", "LINE": "Line"})
                st.dataframe(match_df, use_container_width=True, hide_index=True)
            else:
                st.warning(f"No matches found for `{lookup}`")

    # ═══════════════════════════════════════════
    # TABS
    # ═══════════════════════════════════════════
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Model Wise Float",
        "🔋 Engine & Battery",
        "🔌 Wiring Harness",
        "🎛️ Cockpit Assembly",
        "📈 VIN vs Float",
        "🔄 File Status & Update",
    ])

    # ─────────────────────────────────────────
    # TAB 1 — MODEL WISE FLOAT SUMMARY
    # ─────────────────────────────────────────
    with tab1:
        st.markdown('<div class="section-title">Model Wise Float Summary</div>', unsafe_allow_html=True)
        if not bom_loaded:
            st.info("💡 **BOM details.xlsx** is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        elif not paint_loaded:
            st.info("💡 **Paint Float** report data is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        else:
            # KPI row
            pf_filtered = paint_df if line_filter == "All Lines" else paint_df[paint_df["LINE"] == line_filter]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total Float", f"{int(pf_filtered['TOTAL_FLOAT'].sum()):,}")
            c2.metric("PBS Float", f"{int(pf_filtered['PBS_FLOAT'].sum()):,}")
            c3.metric("Upto Sealant", f"{int(pf_filtered['TOTAL_UPTO_SEALANT'].sum()):,}")
            c4.metric("BIW→PT", f"{int(pf_filtered['BIW_LIFTING_TO_PT'].sum()):,}")

            st.markdown("")

            model_df = compute_model_wise_float(paint_df, dpt_all, line_filter)
            if not model_df.empty:


                # Render using custom HTML table to wrap headers and fit container width without horizontal scroll
                html_table = render_html_table(model_df)
                st.markdown(html_table, unsafe_allow_html=True)
                st.markdown("")

                st.download_button(
                    "📥  Download Model Wise Float",
                    to_excel(model_df, "Model Wise Float", table_type="generic"),
                    file_name=f"model_wise_float_{get_ist_now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.info("No data available for the selected filter.")

    # ─────────────────────────────────────────
    # TAB 2 — ENGINE & BATTERY SUMMARY
    # ─────────────────────────────────────────
    with tab2:
        st.markdown('<div class="section-title">Engine & Battery Summary — Manual Stock Entry</div>', unsafe_allow_html=True)
        st.caption("Enter **Clearance After 6:30AM** values below. Your entries are saved automatically.")

        if not bom_loaded:
            st.info("💡 **BOM details.xlsx** is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        elif not paint_loaded:
            st.info("💡 **Paint Float** report data is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        else:
            engine_req = build_engine_table(paint_df, line_filter)

            if engine_req.empty:
                st.info("No engine data available for the selected filter.")
            else:
                # Load saved engine stock
                saved = load_engine_json()

                # Sum DPT VIN for each engine part number if DPT plans are uploaded
                today_engine_vin = {}
                if dpt_all is not None and not dpt_all.empty:
                    dpt = dpt_all.copy()
                    dpt["ENGINE"] = dpt["ENGINE"].astype(str).str.strip()
                    dpt_agg = dpt[dpt["ENGINE"].str.len() >= 5].groupby("ENGINE", as_index=False)["DPT_VIN"].sum()
                    today_engine_vin = dict(zip(dpt_agg["ENGINE"], dpt_agg["DPT_VIN"]))

                # Build clearance column from saved data
                engine_req["Clearance After 6:30AM"] = engine_req["Engine Part No"].map(lambda e: saved.get(e, {}).get("clearance", 0))
                engine_req["Today VIN"] = engine_req["Engine Part No"].map(
                    lambda e: today_engine_vin.get(e, saved.get(e, {}).get("today_vin", 0))
                )

                # Manual Entry Editor (only real engine rows, no subtotals)
                st.markdown("#### ✏️ Manual Clearance Entry")

                # Sort engines in the desired order (TCF1 Punch, Nova, TCF2 Diesel, EV)
                tcf1_order = ["Punch MT SA", "Punch AMT SA", "Punch TC MCE", "Punch MCE MT", "Punch MCE AMT", "Punch MCE CNG MT", "Punch MCE CNG AMT"]
                tcf2_order = ["Harrier / Safari Diesel AT", "Harrier / Safari Diesel MT", "Harrier / Safari Petrol TGDI MT", "Harrier / Safari Petrol TGDI AT"]
                full_order = tcf1_order + ["Nova"] + tcf2_order + ["Harrier EV"]

                editor_df = engine_req[["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM"]].copy()
                editor_df["_sort"] = editor_df["Model"].map(lambda m: full_order.index(m) if m in full_order else 999)
                editor_df = editor_df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

                # Initialize or update the editor dataframe in session state for stability
                should_init = False
                if "engine_editor_df" not in st.session_state:
                    should_init = True
                else:
                    current_df = st.session_state["engine_editor_df"]
                    if len(current_df) != len(editor_df):
                        should_init = True
                    elif not (current_df["Engine Part No"].values == editor_df["Engine Part No"].values).all():
                        should_init = True

                if should_init:
                    st.session_state["engine_editor_df"] = editor_df

                # Render button to open modal pop-up
                st.markdown("#### ✏️ Manual Clearance Entry")
                if st.button("📝 Open Manual Entry Sheet", type="primary", help="Click to open the manual entry sheet in a pop-up window."):
                    open_manual_entry_popup(st.session_state["engine_editor_df"])

                # Computed Shortage Table
                st.markdown('<div class="section-title">Computed Shortage</div>', unsafe_allow_html=True)

                result = engine_req[[
                    "Engine Part No", "Model", "TA Code", "Clearance After 6:30AM", "LINE"
                ]].copy()
                result["Today VIN"] = engine_req["Today VIN"]
                result["Bal"] = result["Clearance After 6:30AM"] - result["Today VIN"]
                result["PBS FLOAT"] = engine_req["PBS_FLOAT"]
                result["Float UPTO SEALANT"] = engine_req["UPTO_SEALANT"]
                result["TOTAL FLOAT"] = engine_req["TOTAL_FLOAT"]
                result["With respect to PBS FLOAT"] = result["Bal"] - result["PBS FLOAT"]
                result["With respect to Sealant FLOAT"] = result["Bal"] - result["Float UPTO SEALANT"]
                result["With respect to Total FLOAT"] = result["Bal"] - result["TOTAL FLOAT"]

                result = result[[
                    "Engine Part No", "Model", "TA Code", "Clearance After 6:30AM", "Today VIN", "Bal",
                    "PBS FLOAT", "Float UPTO SEALANT", "TOTAL FLOAT",
                    "With respect to PBS FLOAT", "With respect to Sealant FLOAT", "With respect to Total FLOAT", "LINE"
                ]]

                result = add_engine_subtotals(result)

                # KPI for engine (exclude subtotal rows from count and sum)
                real_result = result[result["Engine Part No"].notna() & (result["Engine Part No"] != "")]
                total_short = real_result["With respect to Total FLOAT"].sum()
                crit = (real_result["With respect to Total FLOAT"] < 0).sum()
                ec2, ec3 = st.columns(2)
                ec2.metric("Critical Shortages", crit, delta=f"-{crit}" if crit else "0", delta_color="inverse")
                ec3.metric("Net Balance (Total)", int(total_short))
                st.markdown(render_engine_html_table(result), unsafe_allow_html=True)
                st.markdown("")

                st.download_button(
                    "📥  Download Engine Summary",
                    to_excel(result, "Engine Summary", table_type="engine"),
                    file_name=f"engine_summary_{get_ist_now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ─────────────────────────────────────────
    # TAB 3 — WIRING HARNESS TRACK
    # ─────────────────────────────────────────
    with tab3:
        st.markdown('<div class="section-title">Wiring Harness Shortage Track</div>', unsafe_allow_html=True)

        if not bom_loaded:
            st.info("💡 **BOM details.xlsx** is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        elif not paint_loaded:
            st.info("💡 **Paint Float** report data is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        else:
            wiring_summary = compute_wiring_summary(
                paint_df, wiring_tcf1, wiring_tcf2, pm, dpt_all, line_filter
            )

            if wiring_summary.empty:
                st.info("No wiring data available for the selected filter.")
            else:
                # KPIs
                total_parts = len(wiring_summary)
                critical = (wiring_summary["Shortage for TOTAL FLOAT"] < 0).sum()
                marginal = (wiring_summary["Shortage for TOTAL FLOAT"] == 0).sum()
                healthy = (wiring_summary["Shortage for TOTAL FLOAT"] > 0).sum()

                wc1, wc2, wc3, wc4 = st.columns(4)
                wc1.metric("Total Parts", total_parts)
                wc2.metric("🔴 Critical", critical, delta=f"-{critical}" if critical else "0", delta_color="inverse")
                wc3.metric("🟡 Marginal", marginal)
                wc4.metric("🟢 Healthy", healthy)

                st.markdown("")

                # Shortage Filter Checkbox
                show_only_short = st.checkbox("🔴 Show only parts with shortages (Shortage < 0)", key="wiring_shortage_filter")
                display_df = wiring_summary.copy()
                if show_only_short:
                    display_df = display_df[display_df["Shortage for TOTAL FLOAT"] < 0]
                st.markdown(render_shortage_html_table(display_df, "wiring"), unsafe_allow_html=True)
                st.markdown("")

                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        "📥  Download Excel",
                        to_excel(display_df, "Wiring Summary", table_type="wiring"),
                        file_name=f"wiring_summary_{get_ist_now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with col_dl2:
                    st.download_button(
                        "🖼️  Download Image (PNG)",
                        export_df_to_image(display_df, table_type="wiring"),
                        file_name=f"wiring_summary_{get_ist_now().strftime('%Y%m%d')}.png",
                        mime="image/png",
                        use_container_width=True,
                    )

    # ─────────────────────────────────────────
    # TAB 4 — COCKPIT ASSEMBLY SUMMARY
    # ─────────────────────────────────────────
    with tab4:
        st.markdown('<div class="section-title">Cockpit Assembly Shortage Report</div>', unsafe_allow_html=True)

        if not bom_loaded:
            st.info("💡 **BOM details.xlsx** is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        elif not paint_loaded:
            st.info("💡 **Paint Float** report data is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        else:
            cockpit_summary = compute_cockpit_summary(
                paint_df, cockpit_dfs, pm, dpt_all, line_filter
            )

            if cockpit_summary.empty:
                st.info("No cockpit data available for the selected filter.")
            else:
                # KPIs
                total_cp = len(cockpit_summary)
                crit_cp = (cockpit_summary["Shortage TOTAL FLOAT"] < 0).sum()
                marg_cp = (cockpit_summary["Shortage TOTAL FLOAT"] == 0).sum()
                ok_cp = (cockpit_summary["Shortage TOTAL FLOAT"] > 0).sum()

                cc1, cc2, cc3, cc4 = st.columns(4)
                cc1.metric("Total Parts", total_cp)
                cc2.metric("🔴 Critical", crit_cp, delta=f"-{crit_cp}" if crit_cp else "0", delta_color="inverse")
                cc3.metric("🟡 Marginal", marg_cp)
                cc4.metric("🟢 Healthy", ok_cp)

                st.markdown("")

                # Shortage Filter Checkbox
                show_only_short = st.checkbox("🔴 Show only parts with shortages (Shortage < 0)", key="cockpit_shortage_filter")
                display_df = cockpit_summary.copy()
                if show_only_short:
                    display_df = display_df[display_df["Shortage TOTAL FLOAT"] < 0]
                st.markdown(render_shortage_html_table(display_df, "cockpit"), unsafe_allow_html=True)
                st.markdown("")

                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        "📥  Download Excel",
                        to_excel(display_df, "Cockpit Summary", table_type="cockpit"),
                        file_name=f"cockpit_summary_{get_ist_now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with col_dl2:
                    st.download_button(
                        "🖼️  Download Image (PNG)",
                        export_df_to_image(display_df, table_type="cockpit"),
                        file_name=f"cockpit_summary_{get_ist_now().strftime('%Y%m%d')}.png",
                        mime="image/png",
                        use_container_width=True,
                    )

    # ─────────────────────────────────────────
    # TAB 5 — VIN vs FLOAT SHORTAGE
    # ─────────────────────────────────────────
    with tab5:
        st.markdown('<div class="section-title">VIN Generation vs Paint Float Shortage</div>', unsafe_allow_html=True)
        st.caption(
            "Today VIN (DPT) vs Paint Float stages. "
            "**Negative shortage** = VIN plan exceeds float → potential gap."
        )

        vin_float_df = compute_vin_vs_float(paint_df, dpt_all, line_filter)

        if not bom_loaded:
            st.info("💡 **BOM details.xlsx** is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        elif not paint_loaded:
            st.info("💡 **Paint Float** report data is missing or empty. Please go to the **🔄 File Status & Update** tab to upload it.")
        else:
            vin_float_df = compute_vin_vs_float(paint_df, dpt_all, line_filter)

            if vin_float_df.empty:
                st.info("Upload DPT Plan files to populate VIN generation data.")
            else:
                # KPIs
                total_vc     = len(vin_float_df)
                total_plan   = int(vin_float_df["DPT Plan"].sum())
                total_vin    = int(vin_float_df["Today VIN (DPT)"].sum())
                crit_total   = (vin_float_df["Shortage vs TOTAL Float"] < 0).sum()
                crit_pbs     = (vin_float_df["Shortage vs PBS Float"] < 0).sum()
                crit_seal    = (vin_float_df["Shortage vs Upto Sealant"] < 0).sum()

                k2, k3, k4, k5, k6 = st.columns(5)
                k2.metric("DPT Plan",      f"{total_plan:,}")
                k3.metric("Today VIN",     f"{total_vin:,}")
                k4.metric("🔴 vs PBS",     crit_pbs,   delta=f"-{crit_pbs}"   if crit_pbs   else "0", delta_color="inverse")
                k5.metric("🔴 vs Sealant", crit_seal,  delta=f"-{crit_seal}"  if crit_seal  else "0", delta_color="inverse")
                k6.metric("🔴 vs Total",   crit_total, delta=f"-{crit_total}" if crit_total else "0", delta_color="inverse")

                st.markdown("")

                # Colour-code shortage columns
                def _style_vin_float(df):
                    styled = df.style
                    shortage_cols = ["Shortage vs PBS Float", "Shortage vs Upto Sealant", "Shortage vs TOTAL Float"]

                    # Set custom headers
                    styles = {}
                    for col in df.columns:
                        styles[col] = [{'selector': 'th', 'props': [('background-color', '#334155'), ('color', 'white'), ('font-weight', '700'), ('border', '1px solid #e2e8f0'), ('padding', '10px 14px')]}]
                    styled = styled.set_table_styles(styles, overwrite=False)

                    def _color(val):
                        try:
                            v = float(val)
                            if v < 0:  return "background-color:#fee2e2; color:#991b1b; font-weight:700; border: 1px solid #fecaca;"
                            if v == 0: return "background-color:#fef3c7; color:#b45309; font-weight:600; border: 1px solid #fde68a;"
                            return "background-color:#dcfce7; color:#166534;"
                        except:
                            return ""

                    for col in shortage_cols:
                        if col in df.columns:
                            styled = getattr(styled, "map", getattr(styled, "applymap", None))(_color, subset=[col])

                    # DPT VIN column highlight
                    if "Today VIN (DPT)" in df.columns:
                        styled = getattr(styled, "map", getattr(styled, "applymap", None))(
                            lambda v: "background-color:#dbeafe; color:#1d4ed8; font-weight:600; border: 1px solid #bfdbfe;",
                            subset=["Today VIN (DPT)"]
                        )
                    return styled

                st.markdown(render_vin_float_html_table(vin_float_df), unsafe_allow_html=True)
                st.markdown("")

                # Model-wise summary table
                st.markdown('<div class="section-title">Model-Wise Summary</div>', unsafe_allow_html=True)
                model_summary = (
                    vin_float_df.groupby(["Line", "Model"], as_index=False)
                    .agg(
                        Variants=("Short VC", "count"),
                        DPT_Plan_Sum=("DPT Plan", "sum"),
                        Today_VIN_Sum=("Today VIN (DPT)", "sum"),
                        Total_Float_Sum=("TOTAL Float", "sum"),
                        PBS_Float_Sum=("PBS Float", "sum"),
                        Upto_Sealant_Sum=("Upto Sealant", "sum"),
                        Short_PBS=("Shortage vs PBS Float", "sum"),
                        Short_Seal=("Shortage vs Upto Sealant", "sum"),
                        Short_Total=("Shortage vs TOTAL Float", "sum"),
                    )
                    .rename(columns={
                        "DPT_Plan_Sum":    "DPT Plan",
                        "Today_VIN_Sum":   "Today VIN",
                        "Total_Float_Sum": "TOTAL Float",
                        "PBS_Float_Sum":   "PBS Float",
                        "Upto_Sealant_Sum":"Upto Sealant",
                        "Short_PBS":       "Shortage vs PBS",
                        "Short_Seal":      "Shortage vs Sealant",
                        "Short_Total":     "Shortage vs Total",
                    })
                )
                st.markdown(render_vin_float_summary_html_table(model_summary), unsafe_allow_html=True)
                st.markdown("")

                st.download_button(
                    "📥  Download VIN vs Float",
                    to_excel(vin_float_df, "VIN vs Float", table_type="generic"),
                    file_name=f"vin_vs_float_{get_ist_now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ─────────────────────────────────────────
    # TAB 6 — FILE STATUS & UPDATE
    # ─────────────────────────────────────────
    with tab6:
        st.markdown('<div class="section-title">File Status & Upload Center</div>', unsafe_allow_html=True)
        st.caption(
            "Manage manual file uploads, monitor scanned files in the project workspace, "
            "and view the system auto-reset status."
        )

        # ── 1. Unified Reset & System Banner ──
        next_reset = now.replace(hour=RESET_HOUR, minute=RESET_MINUTE, second=0, microsecond=0)
        if now >= next_reset:
            next_reset = next_reset + timedelta(days=1)
        mins_to_reset = int((next_reset - now).total_seconds() / 60)

        c_info, c_reset = st.columns([3, 1])
        with c_info:
            st.info(f"⏱️ **Next Auto-Reset**: in **{mins_to_reset} min** (at 6:30 AM). All uploaded data will be cleared.")
        with c_reset:
            if st.button("🔄 Manual Daily Reset", use_container_width=True, help="Clear all manually uploaded files and JSON stocks."):
                _perform_reset()
                now = get_ist_now()
                _save_last_reset(now.strftime("%Y-%m-%d"))
                if "engine_editor_df" in st.session_state:
                    del st.session_state["engine_editor_df"]
                st.success("✅ Reset complete — all uploads and caches cleared!")
                st.cache_data.clear()
                st.rerun()

        # ── 2. Grid of File Status Cards ──
        st.markdown("### 📋 Data Channel Status")
        
        # We group RAW_FILE_TYPES into rows of 3 columns
        rows = [RAW_FILE_TYPES[i:i + 3] for i in range(0, len(RAW_FILE_TYPES), 3)]
        
        for row_files in rows:
            cols = st.columns(3)
            for idx, ft in enumerate(row_files):
                with cols[idx]:
                    src_type, src_path = get_source(ft)
                    
                    if src_type == "uploaded":
                        mtime = datetime.fromtimestamp(os.path.getmtime(src_path), tz=timezone.utc).astimezone(timezone(timedelta(hours=5, minutes=30)))
                        age = (now - mtime).total_seconds()
                        pill_class = "pill-badge-success" if age < 86400 else "pill-badge-danger"
                        status_label = "● Uploaded"
                        ts = mtime.strftime("%d-%b %H:%M")
                        orig = file_meta.get(ft, {}).get("original_name", src_path.name)
                        meta_html = f'<div class="card-meta-line"><b>File:</b> {orig}<br><b>Time:</b> {ts}</div>'
                    elif src_type == "scanned":
                        mtime = datetime.fromtimestamp(os.path.getmtime(src_path), tz=timezone.utc).astimezone(timezone(timedelta(hours=5, minutes=30)))
                        age = (now - mtime).total_seconds()
                        pill_class = "pill-badge-info" if age < 86400 else "pill-badge-danger"
                        name_upper = src_path.name.upper()
                        is_con = ("TCF VIN" in name_upper and "FLOAT" in name_upper and "MAPPING" in name_upper)
                        status_label = "◉ Consolidated" if is_con else "◉ Scanned"
                        ts = mtime.strftime("%d-%b %H:%M")
                        meta_html = f'<div class="card-meta-line"><b>File:</b> {src_path.name}<br><b>Time:</b> {ts}</div>'
                    else:
                        pill_class = "pill-badge-warning"
                        status_label = "○ Missing"
                        meta_html = '<div class="card-meta-line" style="color:#94a3b8">— file not loaded —</div>'

                    # Render card HTML
                    st.markdown(
                        f'<div class="file-status-card">'
                        f'  <div class="card-header-line">'
                        f'    <span class="card-title">{ft}</span>'
                        f'    <span class="pill-badge {pill_class}">{status_label}</span>'
                        f'  </div>'
                        f'  {meta_html}'
                        f'</div>',
                        unsafe_allow_html=True
                    )
        
        st.markdown("")

        # ── 3. BOM & Upload Actions side-by-side ──
        c_bom, c_up = st.columns([1, 1])

        with c_bom:
            st.markdown("### 🗂 BOM Master Status")
            bom_info = get_bom_info()
            if bom_info:
                bom_ts = bom_info.get("loaded_at", "Unknown")
                bom_rows = bom_info.get("row_count", 0)
                bom_src = bom_info.get("source", "Unknown")
                st.markdown(
                    f'<div style="background-color:#f0fdf4; border:1px solid #bbf7d0; border-radius:12px; padding:16px;">'
                    f'<span style="color:#166534; font-weight:bold; font-size:1.05rem;">✅ BOM Loaded & Active</span><br>'
                    f'<div style="margin-top:8px; font-size:0.85rem; color:#166534; line-height:1.5;">'
                    f'<b>Variants:</b> {bom_rows} unique Short VCs<br>'
                    f'<b>Source:</b> {bom_src}<br>'
                    f'<b>Import Time:</b> {bom_ts}'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )
                
                # Render BOM Mapping Warnings if any missing VCs
                if missing_bom_vcs:
                    st.markdown("")
                    st.warning(f"⚠️ **BOM Mapping Warning**: Found **{len(missing_bom_vcs)}** active Short VC(s) in Paint Float / DPT Plan without a matching row in the BOM Details file. This can lead to missing wiring, cockpit, or engine allocations!")
                    with st.expander("Show Missing Short VCs"):
                        st.write(", ".join(missing_bom_vcs))
                else:
                    st.markdown("")
                    st.success("✅ **BOM Integrity OK**: All active Short VCs are mapped in BOM details!")
            else:
                st.markdown(
                    f'<div style="background-color:#fef2f2; border:1px solid #fecaca; border-radius:12px; padding:16px;">'
                    f'<span style="color:#991b1b; font-weight:bold; font-size:1.05rem;">⚠️ BOM Master Missing</span><br>'
                    f'<div style="margin-top:8px; font-size:0.85rem; color:#991b1b;">'
                    f'Please ensure <b>Bom details.xlsx</b> is present in the project folder to enable part mappings.'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

        with c_up:
            st.markdown("### 📁 Upload / Update Data")
            selected_ft = st.selectbox("Select file type to update", RAW_FILE_TYPES, key="tab6_ft_select")
            uploaded = st.file_uploader(
                f"Upload **{selected_ft}** file",
                type=["xlsb", "xlsx", "xls"],
                key=f"tab6_upload_{selected_ft}",
            )
            if uploaded is not None:
                state_key = f"last_processed_{selected_ft}"
                upload_sig = f"{uploaded.name}_{uploaded.size}"
                if st.session_state.get(state_key) != upload_sig:
                    save_upload(uploaded, selected_ft)
                    st.session_state[state_key] = upload_sig
                    st.session_state["show_upload_toast"] = f"✅ {selected_ft} saved successfully!"
                    st.cache_data.clear()
                    if "engine_editor_df" in st.session_state:
                        del st.session_state["engine_editor_df"]
                    st.rerun()



# ═══════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()
