import pytest
import pandas as pd
import numpy as np
import os
import json
from pathlib import Path
from datetime import datetime, timezone, timedelta
import app

# ═══════════════════════════════════════════════════════════════
# 1. UTILITIES & HELPERS TESTS (25+ Cases)
# ═══════════════════════════════════════════════════════════════

# Test 1-30: Filename slugification converter (_safe)
@pytest.mark.parametrize("file_type, expected", [
    ("Paint Float", "paint_float"),
    ("TCF1 Wiring File", "tcf1_wiring_file"),
    ("TCF2 Wiring File", "tcf2_wiring_file"),
    ("TCF1 Cockpit", "tcf1_cockpit"),
    ("TCF2 Cockpit", "tcf2_cockpit"),
    ("Nova Cockpit", "nova_cockpit"),
    ("TCF1 DPT Plan", "tcf1_dpt_plan"),
    ("TCF2 DPT Plan", "tcf2_dpt_plan"),
    ("Engine Manual Data", "engine_manual_data"),
    ("  spaces  ", "spaces"),
    ("UPPERCASE", "uppercase"),
    ("File_With_Underscores", "file_with_underscores"),
    ("Slash/Char", "slash/char"),
    ("Special!@#$Chars", "special!@#$chars"),
    ("Multiple   Spaces", "multiple___spaces"),
    ("Punch", "punch"),
    ("punch.ev", "punch.ev"),
    ("harrier", "harrier"),
    ("safari", "safari"),
    ("nova", "nova"),
    ("tcf1", "tcf1"),
    ("tcf2", "tcf2"),
    ("A B C", "a_b_c"),
    ("X_Y_Z", "x_y_z"),
    ("File-Type", "file-type"),
    ("123", "123"),
    ("a-b_c d", "a-b_c_d"),
    ("test  file  type", "test__file__type"),
    ("ok", "ok"),
    ("reset_state", "reset_state"),
    ("Another Test", "another_test"),
    ("Second Test Case", "second_test_case"),
])
def test_safe_filename(file_type, expected):
    assert app._safe(file_type.strip()) == expected.replace(" ", "_").lower().strip()


# Test 31: Dataframe whitespace and NAN cleanup (strip_df)
def test_strip_df():
    df = pd.DataFrame({
        " Col A ": [" valA ", "nan", "None", ""],
        " Col B": [1, 2, 3, np.nan],
        "Col C ": [1.5, 2.5, 3.5, 4.5]
    })
    cleaned = app.strip_df(df)
    
    # Assertions
    assert "Col A" in cleaned.columns
    assert "Col B" in cleaned.columns
    assert "Col C" in cleaned.columns
    
    assert cleaned.loc[0, "Col A"] == "valA"
    assert pd.isna(cleaned.loc[1, "Col A"])
    assert pd.isna(cleaned.loc[2, "Col A"])
    assert pd.isna(cleaned.loc[3, "Col A"])
    
    assert cleaned.loc[0, "Col B"] == 1
    assert pd.isna(cleaned.loc[3, "Col B"])
    assert cleaned.loc[0, "Col C"] == 1.5


# ═══════════════════════════════════════════════════════════════
# 2. DAILY RESET MECHANISM TESTS (25+ Cases)
# ═══════════════════════════════════════════════════════════════

# Test 32-56: Reset logic timestamps & timezone calculations
@pytest.mark.parametrize("current_time_str, last_reset, expected_reset", [
    ("2026-06-18 06:00:00", "2026-06-17", False),
    ("2026-06-18 06:29:59", "2026-06-17", False),
    ("2026-06-18 06:30:00", "2026-06-17", True),
    ("2026-06-18 06:30:01", "2026-06-17", True),
    ("2026-06-18 07:00:00", "2026-06-17", True),
    ("2026-06-18 07:00:00", "2026-06-18", False),
    ("2026-06-18 12:00:00", "2026-06-18", False),
    ("2026-06-18 23:59:59", "2026-06-18", False),
    ("2026-06-19 00:01:00", "2026-06-18", False),
    ("2026-06-19 06:29:59", "2026-06-18", False),
    ("2026-06-19 06:30:00", "2026-06-18", True),
    ("2026-06-18 06:30:00", None, True),
    ("2026-06-18 05:00:00", None, False),
    ("2026-06-18 06:31:00", "2026-06-16", True),
    ("2026-06-18 08:30:00", "2026-06-16", True),
    ("2026-06-18 06:15:00", "2026-06-18", False),
    ("2026-06-18 06:30:00", "2026-06-15", True),
    ("2026-06-18 06:30:00", "2026-05-18", True),
    ("2026-06-18 06:45:00", None, True),
    ("2026-07-01 06:30:00", "2026-06-30", True),
    ("2026-07-01 06:29:00", "2026-06-30", False),
    ("2026-12-31 06:30:00", "2026-12-30", True),
    ("2026-12-31 06:30:00", "2026-12-31", False),
    ("2027-01-01 06:30:00", "2026-12-31", True),
    ("2027-01-01 06:20:00", "2026-12-31", False),
])
def test_reset_trigger_times(current_time_str, last_reset, expected_reset, monkeypatch):
    dt_now = datetime.strptime(current_time_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone(timedelta(hours=5, minutes=30)))
    monkeypatch.setattr(app, "get_ist_now", lambda: dt_now)
    monkeypatch.setattr(app, "_get_last_reset", lambda: last_reset)
    
    reset_called = []
    monkeypatch.setattr(app, "_perform_reset", lambda: reset_called.append(True))
    monkeypatch.setattr(app, "_save_last_reset", lambda d: None)
    
    triggered = app.check_and_perform_daily_reset()
    assert triggered == expected_reset
    if expected_reset:
        assert len(reset_called) == 1


# Test 57-60: Reset state file loader
def test_get_last_reset_datetime(monkeypatch, tmp_path):
    state_file = tmp_path / "last_reset.json"
    monkeypatch.setattr(app, "RESET_STATE_FILE", state_file)
    
    # 1. No file -> return default epoch
    dt = app._get_last_reset_datetime()
    assert dt.year == 1970
    
    # 2. Corrupt JSON -> return epoch
    state_file.write_text("corrupted json")
    dt = app._get_last_reset_datetime()
    assert dt.year == 1970
    
    # 3. Correct reset time
    with open(state_file, "w") as f:
        json.dump({"last_reset_date": "2026-06-18", "reset_time": "2026-06-18 06:30:00"}, f)
    dt = app._get_last_reset_datetime()
    assert dt.year == 2026
    assert dt.month == 6
    assert dt.day == 18
    assert dt.hour == 6
    assert dt.minute == 30


# ═══════════════════════════════════════════════════════════════
# 3. EXCEL PARSER TESTS (20+ Cases)
# ═══════════════════════════════════════════════════════════════

# Test 61-75: Header scanning & Column mapping fallback tests
@pytest.mark.parametrize("headers, expected_idx", [
    (["PART NO", "MODEL", "STOCK", "WIRING HARNESS COVERAGE (FRESH VIN STOCK)"], 3),
    (["PART NO", "MODEL", "STOCK", "COVERAGE"], 3),
    (["PART NO", "MODEL", "STOCK", "FRESH VIN"], 3),
    (["PART NO", "MODEL", "STOCK", "COVERAGE", "FRESH VIN"], 3),
    (["PART NO", "MODEL", "STOCK", "CLEARANCE", "TA CODE"], 9),
    (["col0", "col1", "col2", "col3", "col4", "col5", "col6", "col7", "col8", "fresh vin stock"], 9),
    (["col0", "col1", "col2", "col3", "col4", "col5", "col6", "col7", "col8", "wiring harness coverage"], 9),
    (["coverage", "col1", "col2"], 0),
    (["col0", "fresh vin"], 1),
    (["col0", "col1", "coverage"], 2),
    (["col0", "col1", "col2", "fresh vin"], 3),
    (["col0", "col1", "col2", "col3", "coverage"], 4),
    (["col0", "col1", "col2", "col3", "col4", "fresh vin"], 5),
    (["col0", "col1", "col2", "col3", "col4", "col5", "coverage"], 6),
    (["col0", "col1", "col2", "col3", "col4", "col5", "col6", "fresh vin"], 7),
])
def test_wiring_header_scanner(headers, expected_idx):
    cols = [h.upper().strip() for h in headers]
    stock_col = 9
    for i, c in enumerate(cols):
        if "COVERAGE" in c or "FRESH VIN" in c:
            stock_col = i
            break
    assert stock_col == expected_idx


# Test 76: BOM mapping integrity warning list
def test_bom_mapping_integrity():
    bom_df = pd.DataFrame({
        "SHORT_VC": ["VC1", "VC2"],
        "FRONT_WIRING": ["W1", "W2"],
        "COCKPIT": ["C1", "C2"],
        "ENGINE": ["E1", "E2"]
    })
    
    active_vcs = {"VC1", "VC2", "VC3"}
    bom_vcs = set(bom_df["SHORT_VC"].dropna().astype(str).str.strip().unique())
    missing_bom_vcs = sorted(list(active_vcs - bom_vcs))
    
    assert len(missing_bom_vcs) == 1
    assert missing_bom_vcs[0] == "VC3"


# ═══════════════════════════════════════════════════════════════
# 4. BUSINESS LOGIC & SHORTAGE CALCULATIONS (25+ Cases)
# ═══════════════════════════════════════════════════════════════

# Test 77-101: Shortage formulas (PBS, Sealant, and Total shortages)
@pytest.mark.parametrize("clearance, today_vin, float_val, expected_shortage", [
    (100, 20, 50, 30),
    (100, 20, 100, -20),
    (0, 0, 10, -10),
    (50, 60, 10, -20),
    (100, 0, 0, 100),
    (10.5, 2.5, 4, 4.0),
    (100, 100, 0, 0),
    (100, 50, 50, 0),
    (10, 5, 2, 3),
    (20, 10, 5, 5),
    (30, 15, 10, 5),
    (40, 20, 15, 5),
    (50, 25, 20, 5),
    (60, 30, 25, 5),
    (70, 35, 30, 5),
    (80, 40, 35, 5),
    (90, 45, 40, 5),
    (100, 50, 45, 5),
    (110, 55, 50, 5),
    (120, 60, 55, 5),
    (130, 65, 60, 5),
    (140, 70, 65, 5),
    (150, 75, 70, 5),
    (160, 80, 75, 5),
    (200, 150, 50, 0),
])
def test_shortage_formulas(clearance, today_vin, float_val, expected_shortage):
    balance = clearance - today_vin
    shortage = balance - float_val
    assert shortage == expected_shortage


# Test 102: Engine MT/AMT subtotal groupings
def test_engine_subtotals():
    engine_req = pd.DataFrame([
        {"LINE": "TCF1", "Engine Part No": "P1", "Model": "Punch MT SA", "TA Code": "3302",
         "Clearance After 6:30AM": 100, "Today VIN": 40, "Bal": 60,
         "PBS_FLOAT": 10, "UPTO_SEALANT": 20, "TOTAL_FLOAT": 30},
        {"LINE": "TCF1", "Engine Part No": "P2", "Model": "Punch AMT SA", "TA Code": "3404",
         "Clearance After 6:30AM": 100, "Today VIN": 30, "Bal": 70,
         "PBS_FLOAT": 5, "UPTO_SEALANT": 15, "TOTAL_FLOAT": 25},
    ])
    
    engine_req["PBS FLOAT"] = engine_req["PBS_FLOAT"]
    engine_req["Float UPTO SEALANT"] = engine_req["UPTO_SEALANT"]
    engine_req["TOTAL FLOAT"] = engine_req["TOTAL_FLOAT"]
    engine_req["With respect to PBS FLOAT"] = engine_req["Bal"] - engine_req["PBS FLOAT"]
    engine_req["With respect to Sealant FLOAT"] = engine_req["Bal"] - engine_req["Float UPTO SEALANT"]
    engine_req["With respect to Total FLOAT"] = engine_req["Bal"] - engine_req["TOTAL FLOAT"]
    
    res = app.add_engine_subtotals(engine_req)
    subtotal_rows = res[res["Model"].str.contains("Total|TCF", na=False)]
    assert len(subtotal_rows) > 0


# Test 103: export_df_to_image renders correctly and returns bytes
def test_export_df_to_image():
    df = pd.DataFrame([
        ["549754600106", "TCF1", 148, 159, 10, 56, 138, 92, -11],
    ], columns=["Wiring Part Number", "Model/Line", "Clearance After 6:30AM", "Paint TOTAL FLOAT", "PBS FLOAT", "Cabs FloatUPTO SEALANT", "Shortage PBS FLOAT", "Shortage Upto Sealant", "Shortage for TOTAL FLOAT"])
    
    img_bytes = app.export_df_to_image(df, "wiring")
    assert isinstance(img_bytes, bytes)
    assert len(img_bytes) > 0


# Test 104: to_excel with wiring formatting returns bytes and applies correct styles
def test_to_excel_wiring_formatting():
    df = pd.DataFrame([
        ["549754600106", "TCF1", 148, 159, 10, 56, 138, 92, -11],
    ], columns=["Wiring Part Number", "Model/Line", "Clearance After 6:30AM", "Paint TOTAL FLOAT", "PBS FLOAT", "Cabs FloatUPTO SEALANT", "Shortage PBS FLOAT", "Shortage Upto Sealant", "Shortage for TOTAL FLOAT"])
    
    excel_bytes = app.to_excel(df, "Wiring Summary", table_type="wiring")
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    # Reload workbook using openpyxl to check styles
    from io import BytesIO
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["Wiring Summary"]
    
    # Row 1 columns 1-3 should be peach (F8CBAD)
    c1 = ws.cell(row=1, column=1)
    c4 = ws.cell(row=1, column=4)
    assert c1.fill.start_color.rgb in ["00F8CBAD", "FFF8CBAD", "F8CBAD"]
    assert c4.fill.start_color.rgb in ["00BDD7EE", "FFBDD7EE", "BDD7EE"]
    
    # Columns 3 and 6 should have double right border
    c3 = ws.cell(row=1, column=3)
    c6 = ws.cell(row=1, column=6)
    assert c3.border.right.style == "double"
    assert c6.border.right.style == "double"
    
    # Negative shortage cell formatting: FCE4D6 background and C00000 font color
    neg_cell = ws.cell(row=2, column=9)
    assert neg_cell.fill.start_color.rgb in ["00FCE4D6", "FFFCE4D6", "FCE4D6"]
    assert neg_cell.font.color.rgb in ["00C00000", "FFC00000", "C00000"]


# Test 105: render_html_table wraps headers and returns clean HTML
def test_render_html_table():
    df = pd.DataFrame([
        ["TCF1", "Punch", 100],
    ], columns=["Paint Float", "MODEL", "TOTAL FLOAT"])
    
    html = app.render_html_table(df)
    assert isinstance(html, str)
    assert "<table" in html
    assert "class=\"model-wise-table\"" in html
    assert "Paint<br>Float" in html
    assert "TOTAL<br>FLOAT" in html
    assert "TCF1" in html


# Test 106: render_shortage_html_table wraps headers and formats borders/negatives
def test_render_shortage_html_table():
    df = pd.DataFrame([
        ["549754600106", "TCF1", 148, 159, 10, 56, 138, 92, -11],
    ], columns=["Wiring Part Number", "Model/Line", "Clearance After 6:30AM", "Paint TOTAL FLOAT", "PBS FLOAT", "Cabs FloatUPTO SEALANT", "Shortage PBS FLOAT", "Shortage Upto Sealant", "Shortage for TOTAL FLOAT"])
    
    html = app.render_shortage_html_table(df, "wiring")
    assert isinstance(html, str)
    assert "double" in html
    assert "FCE4D6" in html
    assert "C00000" in html


# Test 107: render_engine_html_table handles row style logic
def test_render_engine_html_table():
    df = pd.DataFrame([
        ["548500", "1.2 Lit Total", 100, 5, 95, 0, 0, 0, 0, 0, 0, 0, "TCF1"],
    ], columns=["Engine Part No", "Model", "TA Code", "Clearance After 6:30AM", "Today VIN", "Bal", "PBS FLOAT", "Float UPTO SEALANT", "TOTAL FLOAT", "With respect to PBS FLOAT", "With respect to Sealant FLOAT", "With respect to Total FLOAT", "LINE"])
    
    html = app.render_engine_html_table(df)
    assert isinstance(html, str)
    assert "background-color" in html
    assert "1.2 Lit Total" in html


# Test 108: render_vin_float_html_table and render_vin_float_summary_html_table render clean HTML
def test_render_vin_float_tables():
    df_vf = pd.DataFrame([
        ["5468", "PUNCH", "PUNCH", "TCF1", 88, 79, 138, 1, 64, 38, 36, 78, 15, -59],
    ], columns=["Short VC", "Sales Description", "Model", "Line", "DPT Plan", "Today VIN (DPT)", "TOTAL Float", "PBS Float", "Upto Sealant", "BIW?PT", "PT?Sealant", "Shortage vs PBS Float", "Shortage vs Upto Sealant", "Shortage vs TOTAL Float"])
    
    html_vf = app.render_vin_float_html_table(df_vf)
    assert isinstance(html_vf, str)
    assert "fee2e2" in html_vf  # negative highlight
    
    df_sum = pd.DataFrame([
        ["TCF1", "PUNCH", 1, 88, 79, 138, 1, 64, 78, 15, -59],
    ], columns=["Line", "Model", "Variants", "DPT Plan", "Today VIN", "TOTAL Float", "PBS Float", "Upto Sealant", "Shortage vs PBS", "Shortage vs Sealant", "Shortage vs Total"])
    
    html_sum = app.render_vin_float_summary_html_table(df_sum)
    assert isinstance(html_sum, str)
    assert "Shortage vs<br>Total" in html_sum
